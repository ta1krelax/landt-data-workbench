#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""蓝电数据工作台：解析、整理、绘图和导出 LAND 混合 CSV。

目标：
1. 自动识别蓝电导出的 Cycle / Step / Record 三层数据。
2. 提供文件管理、指标选择、多轴绘图及样式设置。
3. 导出适合 Origin/Excel 的宽表或长表，以及 PNG/SVG/PDF/EPS/TIFF/EMF 图像。

程序仅依赖 tkinter、numpy、matplotlib 和 openpyxl。前三者用于 GUI/绘图，
openpyxl 仅在导出 xlsx 时使用；CSV 导出不依赖 openpyxl。
"""

from __future__ import annotations

import argparse
import csv
import ctypes
import json
import math
import os
import queue
import re
import shutil
import sys
import tempfile
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, Literal, Sequence

import numpy as np

import matplotlib

matplotlib.use("TkAgg")
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter, MaxNLocator

import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, ttk

matplotlib.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
matplotlib.rcParams["axes.unicode_minus"] = False


APP_NAME = "蓝电数据工作台"
APP_VERSION = "1.3.0"

PLOT_MODE_LABELS = {
    "原始时序图": "time",
    "多电池均值＋误差棒": "statistics",
    "多电池原始曲线 Stack": "stack",
}
ERROR_BAR_LABELS = {"标准差 SD": "sd", "标准误 SEM": "sem", "不显示": "none"}


def enable_windows_dpi_awareness() -> None:
    if os.name != "nt":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass


enable_windows_dpi_awareness()


def safe_float(value: Any) -> float | None:
    """Convert a LAND numeric cell into float; return None for blanks and '-'."""
    if value is None:
        return None
    text = str(value).strip().replace("%", "")
    if not text or text == "-":
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def parse_duration(value: str | None) -> float | None:
    """Parse LAND elapsed time (supports hours > 24, d.hh:mm:ss and d-hh:mm:ss)."""
    if not value:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    try:
        day_match = re.fullmatch(r"(\d+)[.-](\d+):(\d+):(\d+(?:\.\d+)?)", text)
        if day_match:
            days, hours, minutes, seconds = day_match.groups()
            return int(days) * 86400 + float(hours) * 3600 + float(minutes) * 60 + float(seconds)
        days = 0
        parts = text.split(":")
        if len(parts) == 3:
            hours, minutes, seconds = parts
        elif len(parts) == 2:
            hours, minutes, seconds = "0", parts[0], parts[1]
        else:
            return float(text)
        return days * 86400 + float(hours) * 3600 + float(minutes) * 60 + float(seconds)
    except (TypeError, ValueError):
        return None


def parse_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    formats = (
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S.%f",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def decode_land_bytes(data: bytes) -> str:
    """Decode common LAND export encodings without silently losing characters."""
    if data.startswith(b"\xef\xbb\xbf"):
        return data[3:].decode("utf-8")
    if data.startswith(b"\xff\xfe"):
        return data[2:].decode("utf-16-le")
    if data.startswith(b"\xfe\xff"):
        return data[2:].decode("utf-16-be")
    for encoding in ("utf-8", "gb18030", "cp936"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def normalize_unit(header: str) -> tuple[str, str | None]:
    if "/" not in header:
        return header.strip(), None
    name, unit = header.rsplit("/", 1)
    return name.strip(), unit.strip()


def capacity_factor(unit: str | None) -> float:
    return 0.001 if unit and unit.lower().replace("μ", "u").replace("µ", "u") == "uah" else 1.0


def energy_factor(unit: str | None) -> float:
    return 0.001 if unit and unit.lower().replace("μ", "u").replace("µ", "u") == "uwh" else 1.0


def power_factor(unit: str | None) -> float:
    return 0.001 if unit and unit.lower().replace("μ", "u").replace("µ", "u") == "uw" else 1.0


@dataclass(frozen=True)
class MetricDef:
    key: str
    label: str
    unit: str
    source: Literal["record", "cycle"]
    category: str
    default_color: str
    default_style: str = "line"

    @property
    def display(self) -> str:
        return f"{self.label} ({self.unit})" if self.unit else self.label


METRICS: tuple[MetricDef, ...] = (
    MetricDef("voltage_v", "电压", "V", "record", "逐时刻基础数据", "#2563EB"),
    MetricDef("current_ma", "电流", "mA", "record", "逐时刻基础数据", "#DC2626"),
    MetricDef("capacity_mah", "当前步骤容量", "mAh", "record", "逐时刻基础数据", "#7C3AED"),
    MetricDef("specific_capacity_mah_g", "当前步骤比容量", "mAh/g", "record", "逐时刻基础数据", "#9333EA"),
    MetricDef("soc_dod_pct", "SOC/DOD", "%", "record", "逐时刻基础数据", "#0891B2"),
    MetricDef("energy_mwh", "当前步骤能量", "mWh", "record", "逐时刻基础数据", "#CA8A04"),
    MetricDef("specific_energy_wh_kg", "当前步骤比能量", "Wh/kg", "record", "逐时刻基础数据", "#A16207"),
    MetricDef("power_mw", "功率", "mW", "record", "逐时刻基础数据", "#EA580C"),
    MetricDef("dq_dv_mah_v", "dQ/dV", "mAh/V", "record", "逐时刻基础数据", "#0F766E"),
    MetricDef("dv_dq_mv_mah", "dV/dQ", "mV/mAh", "record", "逐时刻基础数据", "#0D9488"),
    MetricDef("temperature_c", "环境温度", "°C", "record", "逐时刻基础数据", "#BE123C"),
    MetricDef("acc_charge_mah", "累计充电容量", "mAh", "record", "累计数据", "#15803D"),
    MetricDef("acc_discharge_mah", "累计放电容量", "mAh", "record", "累计数据", "#65A30D"),
    MetricDef("acc_charge_energy_mwh", "累计充电能量", "mWh", "record", "累计数据", "#B45309"),
    MetricDef("acc_discharge_energy_mwh", "累计放电能量", "mWh", "record", "累计数据", "#D97706"),
    MetricDef("coulombic_efficiency_pct", "库伦效率", "%", "cycle", "循环汇总数据", "#111827", "line+point"),
    MetricDef("charge_capacity_mah", "充电容量", "mAh", "cycle", "循环汇总数据", "#16A34A", "line+point"),
    MetricDef("discharge_capacity_mah", "放电容量", "mAh", "cycle", "循环汇总数据", "#84CC16", "line+point"),
    MetricDef("charge_specific_capacity_mah_g", "充电比容量", "mAh/g", "cycle", "循环汇总数据", "#059669", "line+point"),
    MetricDef("discharge_specific_capacity_mah_g", "放电比容量", "mAh/g", "cycle", "循环汇总数据", "#10B981", "line+point"),
    MetricDef("charge_energy_mwh", "充电能量", "mWh", "cycle", "循环汇总数据", "#F59E0B", "line+point"),
    MetricDef("discharge_energy_mwh", "放电能量", "mWh", "cycle", "循环汇总数据", "#F97316", "line+point"),
    MetricDef("energy_efficiency_pct", "能量效率", "%", "cycle", "循环汇总数据", "#334155", "line+point"),
    MetricDef("charge_average_voltage_v", "充电平均电压", "V", "cycle", "循环汇总数据", "#0369A1", "line+point"),
    MetricDef("discharge_average_voltage_v", "放电平均电压", "V", "cycle", "循环汇总数据", "#0284C7", "line+point"),
    MetricDef("charge_median_voltage_v", "充电中值电压", "V", "cycle", "循环汇总数据", "#1D4ED8", "line+point"),
    MetricDef("discharge_median_voltage_v", "放电中值电压", "V", "cycle", "循环汇总数据", "#3B82F6", "line+point"),
    MetricDef("discharge_end_voltage_v", "放电截止电压", "V", "cycle", "循环汇总数据", "#0EA5E9", "line+point"),
    MetricDef("capacity_retention_pct", "容量保持率", "%", "cycle", "循环汇总数据", "#475569", "line+point"),
    MetricDef("charge_period_s", "充电时长", "s", "cycle", "循环汇总数据", "#64748B", "line+point"),
    MetricDef("discharge_period_s", "放电时长", "s", "cycle", "循环汇总数据", "#94A3B8", "line+point"),
    MetricDef("charge_dcir_mohm", "充电 DCIR", "mΩ", "cycle", "循环汇总数据", "#78716C", "line+point"),
    MetricDef("discharge_dcir_mohm", "放电 DCIR", "mΩ", "cycle", "循环汇总数据", "#57534E", "line+point"),
)

METRIC_BY_KEY = {item.key: item for item in METRICS}


@dataclass
class RecordPoint:
    elapsed_s: float
    system_time: datetime | None
    cycle: int
    step: int
    cycle_step: int
    step_status: str
    step_mode: str
    step_phase: str
    values: dict[str, float | None]


@dataclass
class StepSummary:
    cycle: int
    step: int
    cycle_step: int
    mode: str
    phase: str
    start_time: datetime | None
    period_s: float | None
    values: dict[str, float | None]


@dataclass
class CycleSummary:
    cycle: int
    values: dict[str, float | None]


@dataclass
class LandDataset:
    path: Path
    name: str
    records: list[RecordPoint] = field(default_factory=list)
    steps: dict[tuple[int, int], StepSummary] = field(default_factory=dict)
    cycles: dict[int, CycleSummary] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    enabled: bool = True
    _cycle_bounds: dict[int, tuple[float, float]] = field(default_factory=dict, repr=False)

    @property
    def duration_s(self) -> float:
        return self.records[-1].elapsed_s if self.records else 0.0

    @property
    def cycle_count(self) -> int:
        return len(self.cycles)

    def cycle_anchor(self, cycle: int, mode: Literal["end", "middle"] = "end") -> float | None:
        bounds = self._cycle_bounds.get(cycle)
        if not bounds:
            return None
        return bounds[1] if mode == "end" else (bounds[0] + bounds[1]) / 2

    def rebuild_index(self) -> None:
        bounds: dict[int, tuple[float, float]] = {}
        for row in self.records:
            if row.cycle not in bounds:
                bounds[row.cycle] = (row.elapsed_s, row.elapsed_s)
            else:
                low, high = bounds[row.cycle]
                bounds[row.cycle] = (min(low, row.elapsed_s), max(high, row.elapsed_s))
        self._cycle_bounds = bounds

    def cycle_anchors(
        self,
        cycle_start: int | None = None,
        cycle_end: int | None = None,
    ) -> tuple[np.ndarray, np.ndarray]:
        x_values: list[float] = []
        cycle_values: list[int] = []
        for cycle in sorted(self.cycles):
            if cycle_start is not None and cycle < cycle_start:
                continue
            if cycle_end is not None and cycle > cycle_end:
                continue
            bounds = self._cycle_bounds.get(cycle)
            if bounds:
                x_values.append((bounds[0] + bounds[1]) / 2)
                cycle_values.append(cycle)
        return np.asarray(x_values, dtype=float), np.asarray(cycle_values, dtype=int)

    def series(
        self,
        metric_key: str,
        cycle_start: int | None = None,
        cycle_end: int | None = None,
    ) -> tuple[np.ndarray, np.ndarray]:
        metric = METRIC_BY_KEY[metric_key]
        x_values: list[float] = []
        y_values: list[float] = []
        if metric.source == "record":
            for row in self.records:
                if cycle_start is not None and row.cycle < cycle_start:
                    continue
                if cycle_end is not None and row.cycle > cycle_end:
                    continue
                value = row.values.get(metric_key)
                if value is not None and math.isfinite(value):
                    x_values.append(row.elapsed_s)
                    y_values.append(value)
        else:
            for cycle in sorted(self.cycles):
                if cycle_start is not None and cycle < cycle_start:
                    continue
                if cycle_end is not None and cycle > cycle_end:
                    continue
                value = self.cycles[cycle].values.get(metric_key)
                anchor = self.cycle_anchor(cycle)
                if value is not None and anchor is not None and math.isfinite(value):
                    x_values.append(anchor)
                    y_values.append(value)
        return np.asarray(x_values, dtype=float), np.asarray(y_values, dtype=float)


def row_map(headers: Sequence[str], row: Sequence[str]) -> dict[str, str]:
    return {header.strip(): (row[index].strip() if index < len(row) else "") for index, header in enumerate(headers)}


def find_value(mapping: dict[str, str], *prefixes: str) -> tuple[float | None, str | None]:
    for header, raw_value in mapping.items():
        if any(header.lower().startswith(prefix.lower()) for prefix in prefixes):
            _, unit = normalize_unit(header)
            return safe_float(raw_value), unit
    return None, None


def find_number_header(mapping: dict[str, str], *prefixes: str) -> tuple[float | None, str]:
    for header, raw_value in mapping.items():
        if any(header.lower().startswith(prefix.lower()) for prefix in prefixes):
            return safe_float(raw_value), header.lower().replace("μ", "u").replace("µ", "u")
    return None, ""


def converted_by_header(mapping: dict[str, str], prefixes: Sequence[str], micro_token: str) -> float | None:
    value, header = find_number_header(mapping, *prefixes)
    if value is None:
        return None
    return value * (0.001 if micro_token.lower() in header else 1.0)


def converted(mapping: dict[str, str], prefixes: Sequence[str], converter: Callable[[str | None], float]) -> float | None:
    value, unit = find_value(mapping, *prefixes)
    return None if value is None else value * converter(unit)


def parse_cycle_summary(mapping: dict[str, str]) -> CycleSummary | None:
    cycle = safe_float(mapping.get("Cycle"))
    if cycle is None:
        return None
    cap_c = converted(mapping, ("CapC/",), capacity_factor)
    cap_d = converted(mapping, ("CapD/",), capacity_factor)
    efficiency = safe_float(mapping.get("Efficiency/%"))
    if efficiency is None and cap_c not in (None, 0) and cap_d is not None:
        efficiency = cap_d / cap_c * 100

    dcir_c, dcir_c_unit = find_value(mapping, "DCIR_C/")
    dcir_d, dcir_d_unit = find_value(mapping, "DCIR_D/")
    if dcir_c is not None and dcir_c_unit and dcir_c_unit.lower() == "ohm":
        dcir_c *= 1000
    if dcir_d is not None and dcir_d_unit and dcir_d_unit.lower() == "ohm":
        dcir_d *= 1000

    values: dict[str, float | None] = {
        "coulombic_efficiency_pct": efficiency,
        "charge_capacity_mah": cap_c,
        "discharge_capacity_mah": cap_d,
        "charge_specific_capacity_mah_g": safe_float(mapping.get("SpeCapC/mAh/g")),
        "discharge_specific_capacity_mah_g": safe_float(mapping.get("SpeCapD/mAh/g")),
        "charge_energy_mwh": converted(mapping, ("EnergyC/",), energy_factor),
        "discharge_energy_mwh": converted(mapping, ("EnergyD/",), energy_factor),
        "energy_efficiency_pct": safe_float(mapping.get("Egy-Effi/%")),
        "charge_average_voltage_v": safe_float(mapping.get("AveVoltC/V")),
        "discharge_average_voltage_v": safe_float(mapping.get("AveVoltD/V")),
        "charge_median_voltage_v": safe_float(mapping.get("MedVoltC/V")),
        "discharge_median_voltage_v": safe_float(mapping.get("MedVoltD/V")),
        "discharge_end_voltage_v": safe_float(mapping.get("EndVoltD/V")),
        "capacity_retention_pct": safe_float(mapping.get("RetentionD/%")),
        "charge_period_s": parse_duration(mapping.get("Period_C")),
        "discharge_period_s": parse_duration(mapping.get("Period_D")),
        "charge_dcir_mohm": dcir_c,
        "discharge_dcir_mohm": dcir_d,
    }
    return CycleSummary(cycle=int(cycle), values=values)


def parse_step_summary(mapping: dict[str, str]) -> StepSummary | None:
    cycle = safe_float(mapping.get("CycleNo"))
    step = safe_float(mapping.get("Step"))
    if cycle is None or step is None:
        return None
    values = {
        "capacity_mah": converted(mapping, ("Capacity/",), capacity_factor),
        "specific_capacity_mah_g": safe_float(mapping.get("SpeCap/mAh/g")),
        "energy_mwh": converted(mapping, ("Power/",), energy_factor),
        "specific_energy_wh_kg": safe_float(mapping.get("SpeEnergy/Wh/kg")),
        "charge_average_voltage_v": safe_float(mapping.get("AveVolt/V")),
        "charge_median_voltage_v": safe_float(mapping.get("MedVolt/V")),
        "acc_charge_mah": converted(mapping, ("AccuCapC/",), capacity_factor),
        "acc_discharge_mah": converted(mapping, ("AccuCapD/",), capacity_factor),
    }
    return StepSummary(
        cycle=int(cycle),
        step=int(step),
        cycle_step=0,
        mode=mapping.get("Mode", "").strip(),
        phase="",
        start_time=parse_datetime(mapping.get("StartTime")),
        period_s=parse_duration(mapping.get("Period")),
        values=values,
    )


def parse_record(mapping: dict[str, str], step_modes: dict[tuple[int, int], str]) -> RecordPoint | None:
    cycle = safe_float(mapping.get("CycleNo"))
    step = safe_float(mapping.get("StepNo"))
    elapsed = parse_duration(mapping.get("TestTime"))
    if cycle is None or step is None or elapsed is None:
        return None

    current = converted_by_header(mapping, ("Current/",), "/ua")
    capacity = converted_by_header(mapping, ("Capacity/",), "/uah")
    energy = converted_by_header(mapping, ("Energy/",), "/uwh")
    power = converted_by_header(mapping, ("Power/",), "/uw")
    dq_dv = converted_by_header(mapping, ("dQ/dV/",), "/uah/")
    dv_dq, dv_dq_header = find_number_header(mapping, "dV/dQ/")
    if dv_dq is not None and "/uah" in dv_dq_header:
        dv_dq *= 1000

    record_capacity_unit = find_header_unit(mapping, "Capacity/")
    record_energy_unit = find_header_unit(mapping, "Energy/")
    values: dict[str, float | None] = {
        "voltage_v": safe_float(mapping.get("Voltage/V")),
        "current_ma": current,
        "capacity_mah": capacity,
        "specific_capacity_mah_g": safe_float(mapping.get("SpeCap/mAh/g")),
        "soc_dod_pct": safe_float(mapping.get("SOC|DOD/%")),
        "energy_mwh": energy,
        "specific_energy_wh_kg": safe_float(mapping.get("SpeEnergy/Wh/kg")),
        "power_mw": power,
        "dq_dv_mah_v": dq_dv,
        "dv_dq_mv_mah": dv_dq,
        "temperature_c": safe_float(mapping.get("IncubTemp/dC.")),
        "acc_charge_mah": scaled_plain(mapping.get("CapacityC"), capacity_factor(record_capacity_unit)),
        "acc_discharge_mah": scaled_plain(mapping.get("CapacityD"), capacity_factor(record_capacity_unit)),
        "acc_charge_energy_mwh": scaled_plain(mapping.get("EnergyC"), energy_factor(record_energy_unit)),
        "acc_discharge_energy_mwh": scaled_plain(mapping.get("EnergyD"), energy_factor(record_energy_unit)),
    }
    cycle_i, step_i = int(cycle), int(step)
    return RecordPoint(
        elapsed_s=elapsed,
        system_time=parse_datetime(mapping.get("SysTime")),
        cycle=cycle_i,
        step=step_i,
        cycle_step=0,
        step_status=mapping.get("StepStatus", "").strip(),
        step_mode=step_modes.get((cycle_i, step_i), mapping.get("StepStatus", "").strip()),
        step_phase="",
        values=values,
    )


def find_header_unit(mapping: dict[str, str], prefix: str) -> str | None:
    for header in mapping:
        if header.lower().startswith(prefix.lower()):
            return normalize_unit(header)[1]
    return None


def mapping_header_lower(mapping: dict[str, str], prefix: str) -> str:
    for header in mapping:
        if header.lower().startswith(prefix.lower()):
            return header.lower()
    return ""


def scaled_plain(value: Any, factor: float) -> float | None:
    number = safe_float(value)
    return None if number is None else number * factor


def parse_land_file(path_value: str | Path) -> LandDataset:
    path = Path(path_value)
    text = decode_land_bytes(path.read_bytes()).lstrip("\ufeff")
    dataset = LandDataset(path=path, name=path.stem)
    section: Literal["cycle", "step", "record"] | None = None
    headers: list[str] = []
    step_modes: dict[tuple[int, int], str] = {}

    reader = csv.reader(text.splitlines())
    for line_number, row in enumerate(reader, start=1):
        if not row or not row[0].strip():
            continue
        first = row[0].strip().lstrip("\ufeff")
        if first == "Cycle":
            section, headers = "cycle", [cell.strip() for cell in row]
            continue
        if first == "Step":
            section, headers = "step", [cell.strip() for cell in row]
            continue
        if first == "Record":
            section, headers = "record", [cell.strip() for cell in row]
            continue
        if section is None or not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", first):
            continue

        try:
            mapping = row_map(headers, row)
            if section == "cycle":
                summary = parse_cycle_summary(mapping)
                if summary and summary.cycle not in dataset.cycles:
                    dataset.cycles[summary.cycle] = summary
            elif section == "step":
                summary = parse_step_summary(mapping)
                if summary:
                    key = (summary.cycle, summary.step)
                    dataset.steps[key] = summary
                    step_modes[key] = summary.mode
            else:
                record = parse_record(mapping, step_modes)
                if record:
                    dataset.records.append(record)
        except Exception as exc:  # one malformed row should not lose the whole file
            if len(dataset.warnings) < 20:
                dataset.warnings.append(f"第 {line_number} 行：{exc}")

    if not dataset.cycles:
        raise ValueError("未识别到 Cycle 汇总数据；请确认这是蓝电导出的 CSV。")
    if not dataset.records:
        raise ValueError("未识别到 Record 逐时刻数据；请确认导出时包含记录明细。")
    dataset.records.sort(key=lambda item: (item.elapsed_s, item.cycle, item.step))

    # 将蓝电全局 StepNo 映射为每圈内部从 1 开始的步骤号。
    local_step_map: dict[tuple[int, int], int] = {}
    steps_by_cycle: dict[int, list[int]] = {}
    for cycle_number, source_step in dataset.steps:
        steps_by_cycle.setdefault(cycle_number, []).append(source_step)
    step_phase_map: dict[tuple[int, int], str] = {}
    for cycle_number, source_steps in steps_by_cycle.items():
        waiting_index = 0
        for local_step, source_step in enumerate(sorted(set(source_steps)), start=1):
            key = (cycle_number, source_step)
            summary = dataset.steps[key]
            local_step_map[key] = local_step
            summary.cycle_step = local_step
            mode_upper = summary.mode.upper().replace("-", "").replace("_", "")
            if mode_upper in {"R", "REST", "WAIT", "PAUSE"}:
                waiting_index += 1
                phase = "Waiting" if waiting_index == 1 else f"Waiting {waiting_index}"
            elif mode_upper in {"CCD", "DCC", "DC", "DCHG", "DISCHARGE", "CPD", "CPCD"} or "DISCH" in mode_upper:
                phase = "Discharging"
            elif mode_upper in {"CVC", "CCC", "CC", "CV", "CCCV", "CHG", "CHARGE", "CPC"} or "CHARGE" in mode_upper:
                phase = "Charging"
            else:
                phase = summary.mode or "Unknown"
            summary.phase = phase
            step_phase_map[key] = phase

    # 某些文件的记录表头出现在步骤汇总之前，二次补全步骤名称与圈内步骤号。
    for row in dataset.records:
        row.step_mode = step_modes.get((row.cycle, row.step), row.step_mode)
        row.cycle_step = local_step_map.get((row.cycle, row.step), row.step)
        row.step_phase = step_phase_map.get((row.cycle, row.step), row.step_mode)
    dataset.rebuild_index()
    return dataset


def metric_value_for_record(dataset: LandDataset, row: RecordPoint, metric_key: str) -> float | None:
    metric = METRIC_BY_KEY[metric_key]
    if metric.source == "record":
        return row.values.get(metric_key)
    cycle = dataset.cycles.get(row.cycle)
    return cycle.values.get(metric_key) if cycle else None


def wide_headers(metric_keys: Sequence[str]) -> list[str]:
    return [
        "Time (s)",
        "Cycle",
        "Step (within cycle)",
        *[METRIC_BY_KEY[key].display for key in metric_keys],
        "File",
        "Step Phase",
        "Step Mode",
        "Step Status",
        "Source Step No",
        "System Time",
    ]


def cycle_is_selected(cycle: int, cycle_start: int | None, cycle_end: int | None) -> bool:
    return (cycle_start is None or cycle >= cycle_start) and (cycle_end is None or cycle <= cycle_end)


def iter_wide_rows(
    dataset: LandDataset,
    metric_keys: Sequence[str],
    cycle_start: int | None = None,
    cycle_end: int | None = None,
) -> Iterator[list[Any]]:
    for row in dataset.records:
        if not cycle_is_selected(row.cycle, cycle_start, cycle_end):
            continue
        yield [
            row.elapsed_s,
            row.cycle,
            row.cycle_step,
            *[metric_value_for_record(dataset, row, key) for key in metric_keys],
            dataset.path.name,
            row.step_phase,
            row.step_mode,
            row.step_status,
            row.step,
            row.system_time.strftime("%Y-%m-%d %H:%M:%S") if row.system_time else "",
        ]


LONG_HEADERS = [
    "Time (s)",
    "Cycle",
    "Step (within cycle)",
    "Value",
    "Metric",
    "Unit",
    "File",
    "Step Phase",
    "Step Mode",
    "Step Status",
    "Source Step No",
    "System Time",
]


def iter_long_rows(
    dataset: LandDataset,
    metric_keys: Sequence[str],
    cycle_start: int | None = None,
    cycle_end: int | None = None,
) -> Iterator[list[Any]]:
    for row in dataset.records:
        if not cycle_is_selected(row.cycle, cycle_start, cycle_end):
            continue
        for key in metric_keys:
            value = metric_value_for_record(dataset, row, key)
            if value is None:
                continue
            metric = METRIC_BY_KEY[key]
            yield [
                row.elapsed_s,
                row.cycle,
                row.cycle_step,
                value,
                metric.label,
                metric.unit,
                dataset.path.name,
                row.step_phase,
                row.step_mode,
                row.step_status,
                row.step,
                row.system_time.strftime("%Y-%m-%d %H:%M:%S") if row.system_time else "",
            ]


def cycle_export_headers() -> list[str]:
    return ["File", "Cycle", *[metric.display for metric in METRICS if metric.source == "cycle"]]


def iter_cycle_rows(
    dataset: LandDataset,
    cycle_start: int | None = None,
    cycle_end: int | None = None,
) -> Iterator[list[Any]]:
    cycle_metrics = [metric for metric in METRICS if metric.source == "cycle"]
    for cycle_number in sorted(dataset.cycles):
        if not cycle_is_selected(cycle_number, cycle_start, cycle_end):
            continue
        summary = dataset.cycles[cycle_number]
        yield [dataset.path.name, cycle_number, *[summary.values.get(metric.key) for metric in cycle_metrics]]


def step_export_headers() -> list[str]:
    return [
        "File",
        "Cycle",
        "Step (within cycle)",
        "Source Step No",
        "Phase",
        "Mode",
        "Start Time",
        "Period (s)",
        "Capacity (mAh)",
        "Specific Capacity (mAh/g)",
        "Energy (mWh)",
        "Specific Energy (Wh/kg)",
        "Average Voltage (V)",
        "Median Voltage (V)",
        "Accumulated Charge (mAh)",
        "Accumulated Discharge (mAh)",
    ]


def iter_step_rows(
    dataset: LandDataset,
    cycle_start: int | None = None,
    cycle_end: int | None = None,
) -> Iterator[list[Any]]:
    for key in sorted(dataset.steps):
        summary = dataset.steps[key]
        if not cycle_is_selected(summary.cycle, cycle_start, cycle_end):
            continue
        yield [
            dataset.path.name,
            summary.cycle,
            summary.cycle_step,
            summary.step,
            summary.phase,
            summary.mode,
            summary.start_time.strftime("%Y-%m-%d %H:%M:%S") if summary.start_time else "",
            summary.period_s,
            summary.values.get("capacity_mah"),
            summary.values.get("specific_capacity_mah_g"),
            summary.values.get("energy_mwh"),
            summary.values.get("specific_energy_wh_kg"),
            summary.values.get("charge_average_voltage_v"),
            summary.values.get("charge_median_voltage_v"),
            summary.values.get("acc_charge_mah"),
            summary.values.get("acc_discharge_mah"),
        ]


def safe_sheet_name(name: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "Data"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 2
    while candidate.lower() in {item.lower() for item in existing}:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def export_selected_data(
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    output_path: str | Path,
    layout: Literal["wide", "long"] = "wide",
    cycle_start: int | None = None,
    cycle_end: int | None = None,
    progress: Callable[[str], None] | None = None,
) -> None:
    if not datasets:
        raise ValueError("没有可导出的文件。")
    if not metric_keys:
        raise ValueError("请至少选择一个数据指标。")
    if cycle_start is not None and cycle_end is not None and cycle_start > cycle_end:
        raise ValueError("导出起始圈不能大于结束圈。")
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    suffix = output.suffix.lower()
    headers = wide_headers(metric_keys) if layout == "wide" else LONG_HEADERS
    row_factory = iter_wide_rows if layout == "wide" else iter_long_rows

    if suffix == ".csv":
        with output.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for dataset in datasets:
                if progress:
                    progress(f"正在导出 {dataset.path.name}")
                writer.writerows(row_factory(dataset, metric_keys, cycle_start, cycle_end))
        return

    if suffix != ".xlsx":
        raise ValueError("数据导出仅支持 .csv 或 .xlsx。")

    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("导出 xlsx 需要 openpyxl；可改选 CSV，或安装 openpyxl。") from exc

    workbook = Workbook(write_only=True)
    existing_names: set[str] = set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def append_header(sheet: Any, row_headers: Sequence[str]) -> None:
        cells = []
        for value in row_headers:
            cell = WriteOnlyCell(sheet, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cells.append(cell)
        sheet.append(cells)

    index_sheet = workbook.create_sheet("Index")
    existing_names.add("Index")
    append_header(index_sheet, ["File", "Records", "Cycles", "Steps", "Data sheet(s)"])

    excel_row_limit = 900_000
    for dataset in datasets:
        if progress:
            progress(f"正在写入 {dataset.path.name}")
        selected_record_count = sum(
            1 for row in dataset.records if cycle_is_selected(row.cycle, cycle_start, cycle_end)
        )
        selected_cycle_count = sum(
            1 for cycle in dataset.cycles if cycle_is_selected(cycle, cycle_start, cycle_end)
        )
        selected_step_count = sum(
            1 for summary in dataset.steps.values() if cycle_is_selected(summary.cycle, cycle_start, cycle_end)
        )
        estimated_rows = selected_record_count if layout == "wide" else selected_record_count * max(1, len(metric_keys))
        chunk_count = max(1, math.ceil(estimated_rows / excel_row_limit))
        sheet_names: list[str] = []
        iterator = row_factory(dataset, metric_keys, cycle_start, cycle_end)
        for chunk_index in range(chunk_count):
            base = f"{dataset.name}_{layout}"
            if chunk_count > 1:
                base += f"_{chunk_index + 1}"
            sheet_name = safe_sheet_name(base, existing_names)
            sheet_names.append(sheet_name)
            sheet = workbook.create_sheet(sheet_name)
            sheet.freeze_panes = "A2"
            append_header(sheet, headers)
            for _ in range(excel_row_limit):
                try:
                    sheet.append(next(iterator))
                except StopIteration:
                    break
        index_sheet.append(
            [dataset.path.name, selected_record_count, selected_cycle_count, selected_step_count, "; ".join(sheet_names)]
        )

    cycle_sheet = workbook.create_sheet(safe_sheet_name("Cycle Summary", existing_names))
    append_header(cycle_sheet, cycle_export_headers())
    for dataset in datasets:
        for row in iter_cycle_rows(dataset, cycle_start, cycle_end):
            cycle_sheet.append(row)

    step_sheet = workbook.create_sheet(safe_sheet_name("Step Summary", existing_names))
    append_header(step_sheet, step_export_headers())
    for dataset in datasets:
        for row in iter_step_rows(dataset, cycle_start, cycle_end):
            step_sheet.append(row)

    workbook.save(output)


@dataclass
class CycleMetricStatistics:
    cycles: np.ndarray
    mean: np.ndarray
    sd: np.ndarray
    sem: np.ndarray
    n: np.ndarray
    values: np.ndarray


def compute_cycle_statistics(
    datasets: Sequence[LandDataset],
    metric_key: str,
    cycle_start: int,
    cycle_end: int,
) -> CycleMetricStatistics:
    """Aggregate one cycle-summary metric across files, aligned by cycle number."""
    metric = METRIC_BY_KEY[metric_key]
    if metric.source != "cycle":
        raise ValueError(f"{metric.display} 不是逐圈汇总指标，不能用于多电池统计。")
    if cycle_start < 1 or cycle_end < cycle_start:
        raise ValueError("统计圈数范围无效。")

    cycles = np.arange(cycle_start, cycle_end + 1, dtype=int)
    values = np.full((len(datasets), len(cycles)), np.nan, dtype=float)
    for dataset_index, dataset in enumerate(datasets):
        for cycle_index, cycle in enumerate(cycles):
            summary = dataset.cycles.get(int(cycle))
            value = summary.values.get(metric_key) if summary else None
            if value is not None and math.isfinite(value):
                values[dataset_index, cycle_index] = value

    n = np.sum(np.isfinite(values), axis=0).astype(int)
    mean = np.full(len(cycles), np.nan, dtype=float)
    sd = np.full(len(cycles), np.nan, dtype=float)
    sem = np.full(len(cycles), np.nan, dtype=float)
    for index, count in enumerate(n):
        if count == 0:
            continue
        column = values[:, index]
        valid = column[np.isfinite(column)]
        mean[index] = float(np.mean(valid))
        # A single available cell is drawn with a zero-length error bar instead of a blank.
        sd[index] = float(np.std(valid, ddof=1)) if count >= 2 else 0.0
        sem[index] = sd[index] / math.sqrt(count)
    return CycleMetricStatistics(cycles, mean, sd, sem, n, values)


def statistics_wide_headers(datasets: Sequence[LandDataset], metric_keys: Sequence[str]) -> list[str]:
    headers = ["Cycle (X)"]
    for key in metric_keys:
        metric = METRIC_BY_KEY[key]
        unit = f" ({metric.unit})" if metric.unit else ""
        headers.extend(
            [
                f"{metric.label} Mean{unit} (Y)",
                f"{metric.label} SD{unit} (YErr)",
                f"{metric.label} SEM{unit} (YErr)",
                f"{metric.label} n",
                *[f"{metric.label} | {dataset.path.name}{unit}" for dataset in datasets],
            ]
        )
    return headers


def iter_statistics_wide_rows(
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    cycle_start: int,
    cycle_end: int,
) -> Iterator[list[Any]]:
    results = {
        key: compute_cycle_statistics(datasets, key, cycle_start, cycle_end) for key in metric_keys
    }
    for cycle_index, cycle in enumerate(range(cycle_start, cycle_end + 1)):
        if not any(results[key].n[cycle_index] > 0 for key in metric_keys):
            continue
        row: list[Any] = [cycle]
        for key in metric_keys:
            result = results[key]
            row.extend(
                [
                    None if not math.isfinite(result.mean[cycle_index]) else result.mean[cycle_index],
                    None if not math.isfinite(result.sd[cycle_index]) else result.sd[cycle_index],
                    None if not math.isfinite(result.sem[cycle_index]) else result.sem[cycle_index],
                    int(result.n[cycle_index]),
                    *[
                        None if not math.isfinite(value) else value
                        for value in result.values[:, cycle_index]
                    ],
                ]
            )
        yield row


def export_cycle_statistics(
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    output_path: str | Path,
    cycle_start: int = 1,
    cycle_end: int = 20,
    progress: Callable[[str], None] | None = None,
) -> None:
    """Export Origin-ready mean/SD/SEM/n plus every battery's source values."""
    if len(datasets) < 1:
        raise ValueError("没有可用于统计的文件。")
    cycle_metric_keys = [key for key in metric_keys if METRIC_BY_KEY[key].source == "cycle"]
    if not cycle_metric_keys:
        raise ValueError("多电池统计至少需要一个循环汇总指标。")
    if cycle_start < 1 or cycle_end < cycle_start:
        raise ValueError("统计起始圈不能大于结束圈。")

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    headers = statistics_wide_headers(datasets, cycle_metric_keys)
    rows = list(iter_statistics_wide_rows(datasets, cycle_metric_keys, cycle_start, cycle_end))
    if progress:
        progress(f"正在统计 {len(datasets)} 个电池的第 {cycle_start}–{cycle_end} 圈")

    if output.suffix.lower() == ".csv":
        with output.open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerows(rows)
        return
    if output.suffix.lower() != ".xlsx":
        raise ValueError("统计数据导出仅支持 .csv 或 .xlsx。")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("导出 xlsx 需要 openpyxl；可改选 CSV，或安装 openpyxl。") from exc

    workbook = Workbook()
    wide_sheet = workbook.active
    wide_sheet.title = "Origin Statistics"
    wide_sheet.append(headers)
    for row in rows:
        wide_sheet.append(row)
    wide_sheet.freeze_panes = "A2"

    long_sheet = workbook.create_sheet("Long Statistics")
    long_headers = ["Cycle", "Metric", "Unit", "Mean", "SD", "SEM", "n"]
    long_sheet.append(long_headers)
    for key in cycle_metric_keys:
        metric = METRIC_BY_KEY[key]
        result = compute_cycle_statistics(datasets, key, cycle_start, cycle_end)
        for index, cycle in enumerate(result.cycles):
            if result.n[index] == 0:
                continue
            long_sheet.append(
                [
                    int(cycle),
                    metric.label,
                    metric.unit,
                    float(result.mean[index]),
                    float(result.sd[index]),
                    float(result.sem[index]),
                    int(result.n[index]),
                ]
            )
    long_sheet.freeze_panes = "A2"

    info_sheet = workbook.create_sheet("Files")
    info_sheet.append(["Battery", "File", "Cycles", "Records", "Path"])
    for index, dataset in enumerate(datasets, start=1):
        info_sheet.append([index, dataset.path.name, dataset.cycle_count, len(dataset.records), str(dataset.path)])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in (wide_sheet, long_sheet, info_sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(output)


class BITMAPINFOHEADER(ctypes.Structure):
    _fields_ = [
        ("biSize", ctypes.c_uint32),
        ("biWidth", ctypes.c_int32),
        ("biHeight", ctypes.c_int32),
        ("biPlanes", ctypes.c_uint16),
        ("biBitCount", ctypes.c_uint16),
        ("biCompression", ctypes.c_uint32),
        ("biSizeImage", ctypes.c_uint32),
        ("biXPelsPerMeter", ctypes.c_int32),
        ("biYPelsPerMeter", ctypes.c_int32),
        ("biClrUsed", ctypes.c_uint32),
        ("biClrImportant", ctypes.c_uint32),
    ]


class BITMAPINFO(ctypes.Structure):
    _fields_ = [("bmiHeader", BITMAPINFOHEADER), ("bmiColors", ctypes.c_uint32 * 3)]


class RECT(ctypes.Structure):
    _fields_ = [("left", ctypes.c_long), ("top", ctypes.c_long), ("right", ctypes.c_long), ("bottom", ctypes.c_long)]


def save_figure_emf(figure: Figure, output_path: str | Path, dpi: int = 300) -> None:
    """Save a Windows EMF containing a high-resolution plot bitmap.

    Matplotlib has no native EMF backend. This creates a standards-compliant Enhanced
    Metafile via Windows GDI and embeds the rendered plot at the requested DPI.
    SVG/PDF exports remain the recommended fully-vector formats.
    """
    if os.name != "nt":
        raise RuntimeError("EMF 导出仅在 Windows 上可用；可改用 SVG 或 PDF。")
    try:
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("EMF 导出需要 Pillow。") from exc

    with tempfile.TemporaryDirectory(prefix="landt_emf_") as temp_dir:
        png_path = Path(temp_dir) / "plot.png"
        figure.savefig(png_path, format="png", dpi=dpi, bbox_inches="tight", facecolor="white")
        with Image.open(png_path) as image:
            image = image.convert("RGBA")
            width, height = image.size
            raw = image.tobytes("raw", "BGRA")

        frame = RECT(0, 0, round(width / dpi * 25.4 * 100), round(height / dpi * 25.4 * 100))
        bmi = BITMAPINFO()
        bmi.bmiHeader = BITMAPINFOHEADER(
            ctypes.sizeof(BITMAPINFOHEADER),
            width,
            -height,
            1,
            32,
            0,
            width * height * 4,
            round(dpi / 0.0254),
            round(dpi / 0.0254),
            0,
            0,
        )

        gdi32 = ctypes.windll.gdi32
        gdi32.CreateEnhMetaFileW.argtypes = [ctypes.c_void_p, ctypes.c_wchar_p, ctypes.POINTER(RECT), ctypes.c_wchar_p]
        gdi32.CreateEnhMetaFileW.restype = ctypes.c_void_p
        gdi32.StretchDIBits.argtypes = [
            ctypes.c_void_p,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_void_p,
            ctypes.POINTER(BITMAPINFO),
            ctypes.c_uint,
            ctypes.c_ulong,
        ]
        gdi32.StretchDIBits.restype = ctypes.c_int
        gdi32.CloseEnhMetaFile.argtypes = [ctypes.c_void_p]
        gdi32.CloseEnhMetaFile.restype = ctypes.c_void_p
        gdi32.DeleteEnhMetaFile.argtypes = [ctypes.c_void_p]
        gdi32.DeleteEnhMetaFile.restype = ctypes.c_bool

        output = str(Path(output_path).resolve())
        hdc = gdi32.CreateEnhMetaFileW(None, output, ctypes.byref(frame), f"{APP_NAME}\0Plot\0\0")
        if not hdc:
            raise OSError("无法创建 EMF 文件。")
        buffer = ctypes.create_string_buffer(raw)
        result = gdi32.StretchDIBits(
            hdc,
            0,
            0,
            width,
            height,
            0,
            0,
            width,
            height,
            ctypes.cast(buffer, ctypes.c_void_p),
            ctypes.byref(bmi),
            0,
            0x00CC0020,
        )
        hemf = gdi32.CloseEnhMetaFile(hdc)
        if not hemf or result == 0:
            if hemf:
                gdi32.DeleteEnhMetaFile(hemf)
            raise OSError("写入 EMF 图像失败。")
        gdi32.DeleteEnhMetaFile(hemf)


def save_figure(figure: Figure, output_path: str | Path, dpi: int = 300) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".emf":
        save_figure_emf(figure, output, dpi=dpi)
    else:
        figure.savefig(output, dpi=dpi, bbox_inches="tight", facecolor="white")


@dataclass
class CurveStyle:
    color: str
    plot_type: Literal["line", "point", "line+point"] = "line"
    side: Literal["auto", "left", "right"] = "auto"
    offset: int = 0
    axis_visible: bool = True
    line_width: float = 1.4
    marker_size: float = 3.5


@dataclass
class PlotOptions:
    title: str = "LAND electrochemical data"
    time_unit: Literal["s", "min", "h"] = "h"
    cycle_start: int = 1
    cycle_end: int = 5
    plot_mode: Literal["time", "statistics", "stack"] = "time"
    error_bar: Literal["sd", "sem", "none"] = "sd"
    show_individual_cells: bool = True
    show_bottom_time: bool = True
    show_top_cycle: bool = True
    show_grid: bool = True
    show_legend: bool = True
    max_points_per_curve: int = 12_000


def time_factor_and_label(unit: str) -> tuple[float, str]:
    if unit == "s":
        return 1.0, "Time (s)"
    if unit == "min":
        return 60.0, "Time (min)"
    return 3600.0, "Time (h)"


def minmax_decimate(x_values: np.ndarray, y_values: np.ndarray, max_points: int) -> tuple[np.ndarray, np.ndarray]:
    """Fast min/max decimation that keeps local peaks better than simple striding."""
    count = len(x_values)
    if count <= max_points or max_points < 100:
        return x_values, y_values
    bucket_size = max(2, math.ceil(count / max(2, max_points // 2)))
    selected: list[int] = [0]
    for start in range(1, count - 1, bucket_size):
        end = min(count - 1, start + bucket_size)
        chunk = y_values[start:end]
        if len(chunk) == 0:
            continue
        low = start + int(np.nanargmin(chunk))
        high = start + int(np.nanargmax(chunk))
        selected.extend(sorted({low, high}))
    selected.append(count - 1)
    indices = np.asarray(sorted(set(selected)), dtype=int)
    return x_values[indices], y_values[indices]


def resolved_axis_layout(metric_keys: Sequence[str], styles: dict[str, CurveStyle]) -> dict[str, tuple[str, int]]:
    result: dict[str, tuple[str, int]] = {}
    side_counts = {"left": 0, "right": 0}
    for index, key in enumerate(metric_keys):
        style = styles[key]
        side = style.side if style.side != "auto" else ("left" if index % 2 == 0 else "right")
        automatic_offset = side_counts[side] * 58
        offset = style.offset if style.side != "auto" else automatic_offset
        result[key] = (side, max(0, int(offset)))
        side_counts[side] += 1
    return result


def draw_plot(
    figure: Figure,
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    styles: dict[str, CurveStyle],
    options: PlotOptions,
    primary_dataset: LandDataset | None = None,
) -> list[Any]:
    """Draw selected metrics with one independently configurable y-axis per metric."""
    figure.clear()
    base = figure.add_subplot(111)
    base.set_facecolor("#FFFFFF")
    figure.patch.set_facecolor("#F8FAFC")
    if not datasets or not metric_keys:
        base.text(
            0.5,
            0.5,
            "请导入文件，然后点击右侧 ＋ 添加绘图数据",
            ha="center",
            va="center",
            transform=base.transAxes,
            fontsize=14,
            color="#64748B",
        )
        base.set_xticks([])
        base.set_yticks([])
        return [base]

    time_factor, time_label = time_factor_and_label(options.time_unit)
    layout = resolved_axis_layout(metric_keys, styles)
    axes: list[Any] = []
    all_handles: list[Any] = []
    all_labels: list[str] = []
    min_x = math.inf
    max_x = -math.inf

    for metric_index, metric_key in enumerate(metric_keys):
        metric = METRIC_BY_KEY[metric_key]
        style = styles[metric_key]
        axis = base if metric_index == 0 else base.twinx()
        axes.append(axis)
        side, offset = layout[metric_key]

        if side == "left":
            axis.yaxis.set_label_position("left")
            axis.yaxis.tick_left()
            axis.spines["left"].set_position(("outward", offset))
            axis.spines["left"].set_visible(style.axis_visible)
            axis.spines["right"].set_visible(False)
        else:
            axis.yaxis.set_label_position("right")
            axis.yaxis.tick_right()
            axis.spines["right"].set_position(("outward", offset))
            axis.spines["right"].set_visible(style.axis_visible)
            axis.spines["left"].set_visible(False)

        if style.axis_visible:
            axis.set_ylabel(metric.display, color=style.color, fontsize=9, labelpad=7)
            axis.tick_params(axis="y", colors=style.color, labelsize=8)
        else:
            axis.set_ylabel("")
            axis.tick_params(axis="y", left=False, right=False, labelleft=False, labelright=False)

        for dataset_index, dataset in enumerate(datasets):
            x_values, y_values = dataset.series(metric_key, options.cycle_start, options.cycle_end)
            if len(x_values) == 0:
                continue
            if metric.source == "record":
                x_values, y_values = minmax_decimate(x_values, y_values, options.max_points_per_curve)
            x_values = x_values / time_factor
            min_x = min(min_x, float(np.nanmin(x_values)))
            max_x = max(max_x, float(np.nanmax(x_values)))
            marker = "o" if style.plot_type in ("point", "line+point") else None
            linestyle = "None" if style.plot_type == "point" else "-"
            alpha = 0.90 if len(datasets) == 1 else max(0.48, 0.88 - dataset_index * 0.025)
            line, = axis.plot(
                x_values,
                y_values,
                color=style.color,
                linestyle=linestyle,
                marker=marker,
                markersize=style.marker_size,
                linewidth=style.line_width,
                alpha=alpha,
                label=f"{dataset.name} | {metric.label}",
            )
            all_handles.append(line)
            all_labels.append(line.get_label())

    if math.isfinite(min_x) and math.isfinite(max_x):
        span = max_x - min_x
        margin = span * 0.015 if span > 0 else max(1.0, abs(max_x) * 0.01)
        base.set_xlim(left=min_x - margin, right=max_x + margin)
    else:
        base.set_xlim(0, 1)
        base.text(
            0.5,
            0.5,
            f"第 {options.cycle_start}–{options.cycle_end} 圈没有可绘制的数据",
            ha="center",
            va="center",
            transform=base.transAxes,
            fontsize=13,
            color="#64748B",
        )
    base.set_xlabel(time_label if options.show_bottom_time else "", fontsize=10)
    base.tick_params(axis="x", bottom=options.show_bottom_time, labelbottom=options.show_bottom_time, labelsize=8)
    base.xaxis.set_major_locator(MaxNLocator(nbins=10, min_n_ticks=4))
    base.set_title(options.title, fontsize=13, fontweight="bold", color="#0F172A", pad=14)
    base.grid(options.show_grid, color="#CBD5E1", linewidth=0.6, alpha=0.65)
    base.spines["top"].set_visible(False)
    for axis in axes:
        axis.spines["top"].set_visible(False)

    if options.show_top_cycle:
        reference = primary_dataset if primary_dataset in datasets else datasets[0]
        anchor_x, cycles = reference.cycle_anchors(options.cycle_start, options.cycle_end)
        if len(anchor_x):
            step = max(1, math.ceil(len(anchor_x) / 12))
            indices = np.arange(0, len(anchor_x), step, dtype=int)
            if indices[-1] != len(anchor_x) - 1:
                indices = np.append(indices, len(anchor_x) - 1)
            top_axis = base.twiny()
            top_axis.set_xlim(base.get_xlim())
            top_axis.set_xticks(anchor_x[indices] / time_factor)
            top_axis.set_xticklabels([str(value) for value in cycles[indices]], fontsize=8)
            top_axis.set_xlabel(f"Cycle ({reference.name})", fontsize=10, labelpad=7)
            top_axis.spines["top"].set_color("#475569")
            top_axis.tick_params(axis="x", colors="#475569")
            axes.append(top_axis)

    if options.show_legend and all_handles:
        column_count = 1 if len(all_handles) < 7 else 2
        base.legend(
            all_handles,
            all_labels,
            loc="best",
            fontsize=7,
            frameon=True,
            framealpha=0.90,
            ncol=column_count,
        )

    left_axes = sum(1 for key in metric_keys if layout[key][0] == "left")
    right_axes = len(metric_keys) - left_axes
    figure.subplots_adjust(
        left=min(0.42, 0.09 + max(0, left_axes - 1) * 0.07),
        right=max(0.58, 0.96 - max(0, right_axes - 1) * 0.07),
        top=0.86 if options.show_top_cycle else 0.91,
        bottom=0.11,
    )
    return axes


def _empty_plot_message(figure: Figure, message: str) -> list[Any]:
    figure.clear()
    axis = figure.add_subplot(111)
    axis.set_facecolor("#FFFFFF")
    figure.patch.set_facecolor("#F8FAFC")
    axis.text(0.5, 0.5, message, ha="center", va="center", transform=axis.transAxes, fontsize=13, color="#64748B")
    axis.set_xticks([])
    axis.set_yticks([])
    return [axis]


def _cycle_metric_keys(metric_keys: Sequence[str]) -> list[str]:
    return [key for key in metric_keys if METRIC_BY_KEY[key].source == "cycle"]


def _record_metric_keys(metric_keys: Sequence[str]) -> list[str]:
    return [key for key in metric_keys if METRIC_BY_KEY[key].source == "record"]


def draw_cycle_statistics_plot(
    figure: Figure,
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    styles: dict[str, CurveStyle],
    options: PlotOptions,
) -> list[Any]:
    """Draw cycle-aligned mean lines with SD/SEM error bars across batteries."""
    cycle_keys = _cycle_metric_keys(metric_keys)
    if not datasets or not cycle_keys:
        return _empty_plot_message(figure, "多电池统计需要导入文件并添加至少一个循环汇总指标")

    figure.clear()
    figure.patch.set_facecolor("#F8FAFC")
    axes_array = figure.subplots(len(cycle_keys), 1, sharex=True, squeeze=False)
    axes = [axes_array[index, 0] for index in range(len(cycle_keys))]
    error_labels = {"sd": "SD", "sem": "SEM", "none": "无误差棒"}

    for metric_index, (axis, key) in enumerate(zip(axes, cycle_keys)):
        metric = METRIC_BY_KEY[key]
        style = styles[key]
        result = compute_cycle_statistics(datasets, key, options.cycle_start, options.cycle_end)
        valid = result.n > 0
        axis.set_facecolor("#FFFFFF")

        if options.show_individual_cells:
            for dataset_index, dataset in enumerate(datasets):
                cell_values = result.values[dataset_index]
                cell_valid = np.isfinite(cell_values)
                if np.any(cell_valid):
                    axis.plot(
                        result.cycles[cell_valid],
                        cell_values[cell_valid],
                        color=style.color,
                        linewidth=max(0.7, style.line_width * 0.65),
                        alpha=0.20,
                        marker="o" if len(result.cycles[cell_valid]) <= 30 else None,
                        markersize=max(1.5, style.marker_size * 0.55),
                        label="单个电池" if dataset_index == 0 else "_nolegend_",
                    )

        if np.any(valid):
            y_error: np.ndarray | None
            if options.error_bar == "sd":
                y_error = result.sd[valid]
            elif options.error_bar == "sem":
                y_error = result.sem[valid]
            else:
                y_error = None
            marker = "o" if style.plot_type in ("point", "line+point") else None
            linestyle = "None" if style.plot_type == "point" else "-"
            axis.errorbar(
                result.cycles[valid],
                result.mean[valid],
                yerr=y_error,
                color=style.color,
                ecolor=style.color,
                elinewidth=max(0.8, style.line_width * 0.8),
                capsize=3 if y_error is not None else 0,
                capthick=1.0,
                linewidth=max(1.5, style.line_width),
                linestyle=linestyle,
                marker=marker or "o",
                markersize=max(3.0, style.marker_size),
                label=f"Mean ± {error_labels[options.error_bar]}" if y_error is not None else "Mean",
                zorder=5,
            )
            positive_n = result.n[valid]
            n_text = f"n={int(positive_n[0])}" if np.all(positive_n == positive_n[0]) else f"n={int(np.min(positive_n))}–{int(np.max(positive_n))}"
            axis.text(0.99, 0.96, n_text, transform=axis.transAxes, ha="right", va="top", fontsize=8, color="#64748B")
        else:
            axis.text(0.5, 0.5, "该范围没有可用数据", ha="center", va="center", transform=axis.transAxes, color="#64748B")

        axis.set_ylabel(metric.display, color=style.color, fontsize=9)
        axis.tick_params(axis="y", labelsize=8, colors=style.color)
        axis.grid(options.show_grid, color="#CBD5E1", linewidth=0.6, alpha=0.65)
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)
        if options.show_legend:
            axis.legend(loc="best", fontsize=7, framealpha=0.9)
        if metric_index < len(axes) - 1:
            axis.tick_params(axis="x", labelbottom=False)

    axes[-1].set_xlabel("Cycle", fontsize=10)
    axes[-1].set_xlim(options.cycle_start - 0.4, options.cycle_end + 0.4)
    axes[-1].xaxis.set_major_locator(MaxNLocator(integer=True, nbins=min(12, max(4, options.cycle_end - options.cycle_start + 1))))
    figure.suptitle(options.title or "Multi-battery statistics", fontsize=13, fontweight="bold", color="#0F172A", y=0.985)
    figure.subplots_adjust(left=0.13, right=0.96, top=0.92, bottom=0.10, hspace=0.14)
    return axes


def draw_record_stack_plot(
    figure: Figure,
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    styles: dict[str, CurveStyle],
    options: PlotOptions,
) -> list[Any]:
    """Stack one raw time-series panel per battery for direct cell-to-cell comparison."""
    record_keys = _record_metric_keys(metric_keys)
    if not datasets or not record_keys:
        return _empty_plot_message(figure, "Stack 图需要导入文件并添加电压、电流等逐时刻指标")

    figure.clear()
    figure.patch.set_facecolor("#F8FAFC")
    base_array = figure.subplots(len(datasets), 1, sharex=True, squeeze=False)
    base_axes = [base_array[index, 0] for index in range(len(datasets))]
    all_axes: list[Any] = []
    metric_axes: dict[str, list[Any]] = {key: [] for key in record_keys}
    y_bounds: dict[str, list[float]] = {key: [math.inf, -math.inf] for key in record_keys}
    layout = resolved_axis_layout(record_keys, styles)
    time_factor, time_label = time_factor_and_label(options.time_unit)
    max_x = 0.0

    for dataset_index, (base, dataset) in enumerate(zip(base_axes, datasets)):
        base.set_facecolor("#FFFFFF")
        selected_times = [
            row.elapsed_s
            for row in dataset.records
            if options.cycle_start <= row.cycle <= options.cycle_end
        ]
        if not selected_times:
            base.text(0.5, 0.5, "所选圈数没有逐时刻数据", ha="center", va="center", transform=base.transAxes, color="#64748B")
            base.set_ylabel(dataset.name, fontsize=8, color="#334155")
            all_axes.append(base)
            continue
        range_start_s = min(selected_times)

        panel_handles: list[Any] = []
        panel_labels: list[str] = []
        for metric_index, key in enumerate(record_keys):
            metric = METRIC_BY_KEY[key]
            style = styles[key]
            axis = base if metric_index == 0 else base.twinx()
            all_axes.append(axis)
            metric_axes[key].append(axis)
            side, offset = layout[key]
            if side == "left":
                axis.yaxis.set_label_position("left")
                axis.yaxis.tick_left()
                axis.spines["left"].set_position(("outward", offset))
                axis.spines["left"].set_visible(style.axis_visible)
                axis.spines["right"].set_visible(False)
            else:
                axis.yaxis.set_label_position("right")
                axis.yaxis.tick_right()
                axis.spines["right"].set_position(("outward", offset))
                axis.spines["right"].set_visible(style.axis_visible)
                axis.spines["left"].set_visible(False)
            axis.spines["top"].set_visible(False)
            if style.axis_visible:
                axis.set_ylabel(metric.display, color=style.color, fontsize=8, labelpad=6)
                axis.tick_params(axis="y", colors=style.color, labelsize=7)
            else:
                axis.set_ylabel("")
                axis.tick_params(axis="y", left=False, right=False, labelleft=False, labelright=False)

            x_values, y_values = dataset.series(key, options.cycle_start, options.cycle_end)
            if len(x_values) == 0:
                continue
            x_values = (x_values - range_start_s) / time_factor
            x_values, y_values = minmax_decimate(x_values, y_values, options.max_points_per_curve)
            valid = np.isfinite(x_values) & np.isfinite(y_values)
            if not np.any(valid):
                continue
            x_values = x_values[valid]
            y_values = y_values[valid]
            max_x = max(max_x, float(np.nanmax(x_values)))
            y_bounds[key][0] = min(y_bounds[key][0], float(np.nanmin(y_values)))
            y_bounds[key][1] = max(y_bounds[key][1], float(np.nanmax(y_values)))
            marker = "o" if style.plot_type in ("point", "line+point") else None
            linestyle = "None" if style.plot_type == "point" else "-"
            line, = axis.plot(
                x_values,
                y_values,
                color=style.color,
                linestyle=linestyle,
                marker=marker,
                markersize=style.marker_size,
                linewidth=style.line_width,
                alpha=0.92,
                label=metric.display,
            )
            panel_handles.append(line)
            panel_labels.append(metric.display)

        base.text(
            0.008,
            0.96,
            dataset.name,
            transform=base.transAxes,
            ha="left",
            va="top",
            fontsize=8.5,
            fontweight="bold",
            color="#0F172A",
            bbox={"boxstyle": "round,pad=0.22", "facecolor": "white", "edgecolor": "#CBD5E1", "alpha": 0.88},
            zorder=10,
        )
        base.grid(options.show_grid, color="#CBD5E1", linewidth=0.55, alpha=0.60)
        base.xaxis.set_major_locator(MaxNLocator(nbins=10, min_n_ticks=4))
        if dataset_index < len(base_axes) - 1 or not options.show_bottom_time:
            base.tick_params(axis="x", labelbottom=False, bottom=options.show_bottom_time)
        else:
            base.tick_params(axis="x", labelsize=8)

        if options.show_top_cycle:
            anchors, cycles = dataset.cycle_anchors(options.cycle_start, options.cycle_end)
            if len(anchors):
                anchors = (anchors - range_start_s) / time_factor
                label_step = max(1, math.ceil(len(anchors) / 12))
                for marker_index, (anchor, cycle) in enumerate(zip(anchors, cycles)):
                    base.axvline(anchor, color="#94A3B8", linewidth=0.55, linestyle=":", alpha=0.65, zorder=0)
                    if marker_index % label_step == 0 or marker_index == len(anchors) - 1:
                        base.text(
                            anchor,
                            0.98,
                            f"C{cycle}",
                            transform=base.get_xaxis_transform(),
                            ha="center",
                            va="top",
                            fontsize=6.8,
                            color="#64748B",
                        )
        if options.show_legend and dataset_index == 0 and panel_handles:
            base.legend(panel_handles, panel_labels, loc="lower right", fontsize=6.8, ncol=min(3, len(panel_handles)), framealpha=0.88)

    for key, axes_for_metric in metric_axes.items():
        low, high = y_bounds[key]
        if not math.isfinite(low) or not math.isfinite(high):
            continue
        span = high - low
        margin = span * 0.06 if span > 0 else max(abs(high) * 0.04, 0.05)
        for axis in axes_for_metric:
            axis.set_ylim(low - margin, high + margin)

    x_margin = max_x * 0.015 if max_x > 0 else 1.0
    base_axes[-1].set_xlim(-x_margin, max_x + x_margin)
    base_axes[-1].set_xlabel(time_label if options.show_bottom_time else "", fontsize=10)
    left_axes = sum(1 for key in record_keys if layout[key][0] == "left")
    right_axes = len(record_keys) - left_axes
    figure.suptitle(options.title or "Multi-battery raw-curve stack", fontsize=13, fontweight="bold", color="#0F172A", y=0.992)
    figure.subplots_adjust(
        left=min(0.42, 0.10 + max(0, left_axes - 1) * 0.065),
        right=max(0.58, 0.97 - max(0, right_axes - 1) * 0.065),
        top=0.94,
        bottom=0.08,
        hspace=0.12,
    )
    return all_axes


def draw_selected_plot(
    figure: Figure,
    datasets: Sequence[LandDataset],
    metric_keys: Sequence[str],
    styles: dict[str, CurveStyle],
    options: PlotOptions,
    primary_dataset: LandDataset | None = None,
) -> list[Any]:
    if options.plot_mode == "statistics":
        return draw_cycle_statistics_plot(figure, datasets, metric_keys, styles, options)
    if options.plot_mode == "stack":
        return draw_record_stack_plot(figure, datasets, metric_keys, styles, options)
    return draw_plot(figure, datasets, metric_keys, styles, options, primary_dataset)


class CollapsibleFrame(ttk.Frame):
    def __init__(self, master: tk.Misc, title: str, initially_open: bool = True, **kwargs: Any) -> None:
        super().__init__(master, **kwargs)
        self.open_var = tk.BooleanVar(value=initially_open)
        self.header = ttk.Button(self, text="", command=self.toggle, style="Accordion.TButton")
        self.header.pack(fill="x")
        self.body = ttk.Frame(self, style="Card.TFrame", padding=(10, 8))
        self.title = title
        self._sync()

    def toggle(self) -> None:
        self.open_var.set(not self.open_var.get())
        self._sync()

    def _sync(self) -> None:
        is_open = self.open_var.get()
        self.header.configure(text=f"{'▾' if is_open else '▸'}  {self.title}")
        if is_open:
            self.body.pack(fill="x")
        else:
            self.body.pack_forget()


class ScrollableFrame(ttk.Frame):
    def __init__(self, master: tk.Misc, **kwargs: Any) -> None:
        super().__init__(master, **kwargs)
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#F1F5F9")
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.inner = ttk.Frame(self.canvas, style="Panel.TFrame")
        self.window_id = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<Enter>", lambda _event: self.canvas.bind_all("<MouseWheel>", self._on_wheel))
        self.canvas.bind("<Leave>", lambda _event: self.canvas.unbind_all("<MouseWheel>"))

    def _on_inner_configure(self, _event: tk.Event) -> None:
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event: tk.Event) -> None:
        self.canvas.itemconfigure(self.window_id, width=event.width)

    def _on_wheel(self, event: tk.Event) -> None:
        self.canvas.yview_scroll(int(-event.delta / 120), "units")


class AddMetricDialog(tk.Toplevel):
    """Compact + dialog for adding one or more metrics to the plot area."""

    def __init__(
        self,
        master: tk.Misc,
        existing_keys: Sequence[str],
        source_filter: Literal["record", "cycle"] | None = None,
    ) -> None:
        super().__init__(master)
        self.title("＋ 添加绘图数据")
        self.geometry("640x700")
        self.minsize(540, 520)
        self.transient(master)
        self.grab_set()
        self.result: list[str] | None = None
        self.existing_keys = set(existing_keys)

        frame = ttk.Frame(self, padding=22, style="Panel.TFrame")
        frame.pack(fill="both", expand=True)
        heading = "选择循环汇总指标" if source_filter == "cycle" else "选择要加入绘图区的数据"
        ttk.Label(frame, text=heading, style="Section.TLabel").pack(anchor="w")
        ttk.Label(
            frame,
            text="支持 Ctrl / Shift 多选；双击可立即添加。",
            style="Hint.TLabel",
            wraplength=570,
            justify="left",
        ).pack(anchor="w", pady=(6, 14))

        tree_frame = ttk.Frame(frame, style="Panel.TFrame")
        tree_frame.pack(fill="both", expand=True)
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical")
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("state",),
            show="tree headings",
            selectmode="extended",
            yscrollcommand=scrollbar.set,
            style="Metric.Treeview",
        )
        scrollbar.configure(command=self.tree.yview)
        self.tree.heading("#0", text="数据")
        self.tree.heading("state", text="状态")
        self.tree.column("#0", width=440, minwidth=300)
        self.tree.column("state", width=105, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        categories: dict[str, list[MetricDef]] = {}
        for metric in METRICS:
            if source_filter is not None and metric.source != source_filter:
                continue
            categories.setdefault(metric.category, []).append(metric)
        for category, metrics in categories.items():
            parent_id = f"category::{category}"
            self.tree.insert("", "end", iid=parent_id, text=category, values=("",), open=True, tags=("category",))
            for metric in metrics:
                state = "已添加" if metric.key in self.existing_keys else ""
                self.tree.insert(parent_id, "end", iid=metric.key, text=metric.display, values=(state,))
        self.tree.tag_configure("category", font=("Segoe UI", 10, "bold"), foreground="#0F2E4F")

        self.tree.bind("<Double-1>", lambda _event: self._accept())
        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(16, 0))
        ttk.Button(buttons, text="取消", command=self.destroy).pack(side="right")
        ttk.Button(buttons, text="添加", command=self._accept, style="Accent.TButton").pack(side="right", padx=(0, 8))
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _accept(self) -> None:
        keys = [item for item in self.tree.selection() if item in METRIC_BY_KEY and item not in self.existing_keys]
        if not keys:
            messagebox.showinfo(APP_NAME, "请选择至少一个尚未添加的数据指标。", parent=self)
            return
        self.result = keys
        self.destroy()


class DataExportDialog(tk.Toplevel):
    def __init__(
        self,
        master: tk.Misc,
        cycle_start: int = 1,
        cycle_end: int = 5,
        default_layout: Literal["wide", "long", "statistics"] = "wide",
    ) -> None:
        super().__init__(master)
        self.title("导出整理后的数据")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result: dict[str, Any] | None = None
        self.layout_var = tk.StringVar(value=default_layout)
        self.scope_var = tk.StringVar(value="selected")
        self.cycle_start_var = tk.IntVar(value=cycle_start)
        self.cycle_end_var = tk.IntVar(value=cycle_end)

        frame = ttk.Frame(self, padding=20, style="Panel.TFrame")
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="数据布局", style="Section.TLabel").pack(anchor="w")
        ttk.Radiobutton(
            frame,
            text="宽表：Time、Cycle、Step 后每个指标一列",
            variable=self.layout_var,
            value="wide",
        ).pack(anchor="w", pady=(8, 2))
        ttk.Radiobutton(
            frame,
            text="长表：第 4 列为 Value，后接 Metric 和 Unit",
            variable=self.layout_var,
            value="long",
        ).pack(anchor="w", pady=2)
        ttk.Radiobutton(
            frame,
            text="多电池统计表：Mean、SD、SEM、n 与各电池数据（Origin）",
            variable=self.layout_var,
            value="statistics",
        ).pack(anchor="w", pady=2)

        ttk.Separator(frame).pack(fill="x", pady=14)
        ttk.Label(frame, text="指标范围", style="Section.TLabel").pack(anchor="w")
        ttk.Radiobutton(frame, text="仅导出当前绘图区指标", variable=self.scope_var, value="selected").pack(anchor="w", pady=(8, 2))
        ttk.Radiobutton(frame, text="导出全部可用指标", variable=self.scope_var, value="all").pack(anchor="w", pady=2)

        ttk.Separator(frame).pack(fill="x", pady=14)
        ttk.Label(frame, text="导出圈数范围", style="Section.TLabel").pack(anchor="w")
        range_row = ttk.Frame(frame, style="Panel.TFrame")
        range_row.pack(fill="x", pady=(8, 2))
        ttk.Label(range_row, text="从", style="Hint.TLabel").pack(side="left")
        ttk.Spinbox(range_row, from_=1, to=1_000_000, textvariable=self.cycle_start_var, width=9).pack(side="left", padx=(6, 14))
        ttk.Label(range_row, text="到", style="Hint.TLabel").pack(side="left")
        ttk.Spinbox(range_row, from_=1, to=1_000_000, textvariable=self.cycle_end_var, width=9).pack(side="left", padx=6)
        ttk.Label(frame, text="默认沿用当前绘图范围，可在此单独修改。", style="Hint.TLabel").pack(anchor="w")

        ttk.Label(
            frame,
            text="XLSX 会按文件拆分数据表，并附 Cycle Summary 与 Step Summary；CSV 为单表。",
            style="Hint.TLabel",
            wraplength=430,
        ).pack(anchor="w", pady=(14, 10))

        buttons = ttk.Frame(frame, style="Panel.TFrame")
        buttons.pack(fill="x", pady=(8, 0))
        ttk.Button(buttons, text="取消", command=self.destroy).pack(side="right")
        ttk.Button(buttons, text="下一步…", command=self._accept, style="Accent.TButton").pack(side="right", padx=(0, 8))
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.update_idletasks()
        x = master.winfo_rootx() + (master.winfo_width() - self.winfo_width()) // 2
        y = master.winfo_rooty() + (master.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{max(0, x)}+{max(0, y)}")

    def _accept(self) -> None:
        try:
            cycle_start = int(self.cycle_start_var.get())
            cycle_end = int(self.cycle_end_var.get())
        except (tk.TclError, ValueError):
            messagebox.showerror(APP_NAME, "请输入有效的导出圈数。", parent=self)
            return
        if cycle_start < 1 or cycle_end < 1 or cycle_start > cycle_end:
            messagebox.showerror(APP_NAME, "导出圈数需为正整数，且起始圈不能大于结束圈。", parent=self)
            return
        self.result = {
            "layout": self.layout_var.get(),
            "scope": self.scope_var.get(),
            "cycle_start": cycle_start,
            "cycle_end": cycle_end,
        }
        self.destroy()


class LandtWorkbenchApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(f"{APP_NAME}  {APP_VERSION}")
        self.geometry("1560x940")
        self.minsize(1220, 760)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.option_add("*Font", "{Segoe UI} 10")
        self._configure_styles()

        self.datasets: dict[str, LandDataset] = {}
        self.loading_paths: set[str] = set()
        self.executor = ThreadPoolExecutor(max_workers=min(4, max(2, (os.cpu_count() or 2) // 2)))
        self.ui_queue: queue.Queue[tuple[str, Any]] = queue.Queue()
        self.curve_styles = {
            metric.key: CurveStyle(metric.default_color, metric.default_style) for metric in METRICS
        }
        self.plot_metric_keys: list[str] = []

        self.status_var = tk.StringVar(value="就绪")
        self.file_detail_var = tk.StringVar(value="尚未导入文件")
        self.curve_count_var = tk.StringVar(value="尚未添加数据；点击 ＋ 开始")
        self.plot_range_hint_var = tk.StringVar(value="默认只绘制第 1–5 圈")
        self.plot_cycle_start_var = tk.IntVar(value=1)
        self.plot_cycle_end_var = tk.IntVar(value=5)
        self.plot_mode_var = tk.StringVar(value="原始时序图")
        self.error_bar_var = tk.StringVar(value="标准差 SD")
        self.show_individual_cells_var = tk.BooleanVar(value=True)
        self.title_var = tk.StringVar(value="LAND electrochemical data")
        self.time_unit_var = tk.StringVar(value="h")
        self.show_bottom_var = tk.BooleanVar(value=True)
        self.show_top_var = tk.BooleanVar(value=True)
        self.show_grid_var = tk.BooleanVar(value=True)
        self.show_legend_var = tk.BooleanVar(value=True)
        self.max_points_var = tk.IntVar(value=12_000)
        self.dpi_var = tk.IntVar(value=300)

        self.style_metric_var = tk.StringVar()
        self.plot_type_var = tk.StringVar(value="线")
        self.side_var = tk.StringVar(value="自动")
        self.offset_var = tk.IntVar(value=0)
        self.axis_visible_var = tk.BooleanVar(value=True)
        self.line_width_var = tk.DoubleVar(value=1.4)
        self.marker_size_var = tk.DoubleVar(value=3.5)

        self._build_ui()
        self._sync_plot_mode_controls()
        self._refresh_curve_list()
        self._refresh_style_metric_choices()
        self.after(100, self._poll_ui_queue)
        self.after(250, self.draw_current_plot)

    def _configure_styles(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(".", background="#F1F5F9", foreground="#0F172A")
        style.configure("TLabel", padding=(0, 2))
        style.configure("TButton", padding=(10, 6))
        style.configure("TCheckbutton", padding=(0, 3))
        style.configure("Panel.TFrame", background="#F1F5F9")
        style.configure("Card.TFrame", background="#FFFFFF")
        style.configure("Toolbar.TFrame", background="#0F2E4F")
        style.configure("Toolbar.TLabel", background="#0F2E4F", foreground="#FFFFFF", font=("Segoe UI", 15, "bold"))
        style.configure("Section.TLabel", background="#F1F5F9", foreground="#0F172A", font=("Segoe UI", 10, "bold"))
        style.configure("CardSection.TLabel", background="#FFFFFF", foreground="#0F172A", font=("Segoe UI", 10, "bold"))
        style.configure("Hint.TLabel", background="#F1F5F9", foreground="#64748B", font=("Segoe UI", 9))
        style.configure("CardHint.TLabel", background="#FFFFFF", foreground="#64748B", font=("Segoe UI", 9))
        style.configure("Accent.TButton", background="#2563EB", foreground="#FFFFFF", borderwidth=0, padding=(12, 7))
        style.map("Accent.TButton", background=[("active", "#1D4ED8"), ("pressed", "#1E40AF")])
        style.configure("Toolbar.TButton", background="#173F68", foreground="#FFFFFF", borderwidth=0, padding=(11, 7))
        style.map("Toolbar.TButton", background=[("active", "#245786")])
        style.configure("Accordion.TButton", background="#E2E8F0", foreground="#0F2E4F", anchor="w", borderwidth=0, padding=(9, 7), font=("Segoe UI", 10, "bold"))
        style.map("Accordion.TButton", background=[("active", "#CBD5E1")])
        style.configure("Treeview", rowheight=30, background="#FFFFFF", fieldbackground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", background="#DCE6F1", foreground="#0F2E4F", font=("Segoe UI", 9, "bold"))
        style.configure("Metric.Treeview", rowheight=32, background="#FFFFFF", fieldbackground="#FFFFFF", borderwidth=0, font=("Segoe UI", 10))
        style.map("Treeview", background=[("selected", "#DBEAFE")], foreground=[("selected", "#0F172A")])
        style.configure("Status.TLabel", background="#E2E8F0", foreground="#334155", padding=(8, 4))

    def _build_ui(self) -> None:
        toolbar = ttk.Frame(self, style="Toolbar.TFrame", padding=(14, 10))
        toolbar.pack(fill="x")
        ttk.Label(toolbar, text=APP_NAME, style="Toolbar.TLabel").pack(side="left", padx=(0, 22))
        ttk.Button(toolbar, text="导入文件", command=self.add_files, style="Toolbar.TButton").pack(side="left", padx=3)
        ttk.Button(toolbar, text="导入文件夹", command=self.add_folder, style="Toolbar.TButton").pack(side="left", padx=3)
        ttk.Button(toolbar, text="导出数据", command=self.export_data_dialog, style="Toolbar.TButton").pack(side="left", padx=(14, 3))
        ttk.Button(toolbar, text="导出图像", command=self.export_figure_dialog, style="Toolbar.TButton").pack(side="left", padx=3)

        self.progress = ttk.Progressbar(toolbar, mode="indeterminate", length=130)
        self.progress.pack(side="right", padx=12)

        main = ttk.Frame(self, style="Panel.TFrame")
        main.pack(fill="both", expand=True)
        main.columnconfigure(0, minsize=310)
        main.columnconfigure(1, weight=1)
        main.columnconfigure(2, minsize=370)
        main.rowconfigure(0, weight=1)

        file_panel = ttk.Frame(main, style="Panel.TFrame", padding=12, width=310)
        plot_panel = ttk.Frame(main, style="Card.TFrame", padding=4)
        settings_panel = ttk.Frame(main, style="Panel.TFrame", width=370)
        file_panel.grid(row=0, column=0, sticky="nsew")
        plot_panel.grid(row=0, column=1, sticky="nsew")
        settings_panel.grid(row=0, column=2, sticky="nsew")
        file_panel.pack_propagate(False)
        settings_panel.pack_propagate(False)

        self._build_file_panel(file_panel)
        self._build_plot_panel(plot_panel)
        self._build_settings_panel(settings_panel)

        ttk.Label(self, textvariable=self.status_var, style="Status.TLabel", anchor="w").pack(fill="x", side="bottom")

    def _build_file_panel(self, parent: ttk.Frame) -> None:
        ttk.Label(parent, text="文件管理", style="Section.TLabel").pack(anchor="w", pady=(0, 8))
        columns = ("status", "cycles", "records")
        tree_frame = ttk.Frame(parent, style="Panel.TFrame")
        tree_frame.pack(fill="both", expand=True)
        tree_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        self.file_tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            selectmode="extended",
            yscrollcommand=tree_y.set,
            xscrollcommand=tree_x.set,
        )
        tree_y.configure(command=self.file_tree.yview)
        tree_x.configure(command=self.file_tree.xview)
        self.file_tree.heading("#0", text="使用 / 文件")
        self.file_tree.heading("status", text="状态")
        self.file_tree.heading("cycles", text="圈")
        self.file_tree.heading("records", text="记录")
        self.file_tree.column("#0", width=118, minwidth=92)
        self.file_tree.column("status", width=48, anchor="center")
        self.file_tree.column("cycles", width=38, anchor="e")
        self.file_tree.column("records", width=62, anchor="e")
        self.file_tree.grid(row=0, column=0, sticky="nsew")
        tree_y.grid(row=0, column=1, sticky="ns")
        tree_x.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        self.file_tree.tag_configure("error", foreground="#B91C1C")
        self.file_tree.tag_configure("loading", foreground="#64748B")
        self.file_tree.bind("<Button-1>", self._file_tree_click)
        self.file_tree.bind("<<TreeviewSelect>>", self._on_file_selected)
        self.file_tree.bind("<Double-1>", self._show_file_details)

        button_row = ttk.Frame(parent, style="Panel.TFrame")
        button_row.pack(fill="x", pady=(8, 5))
        ttk.Button(button_row, text="切换使用", command=self.toggle_selected_files).pack(side="left", padx=(0, 4))
        ttk.Button(button_row, text="移除", command=self.remove_selected_files).pack(side="left", padx=4)
        ttk.Button(button_row, text="清空", command=self.clear_files).pack(side="right")

        detail_card = ttk.Frame(parent, style="Card.TFrame", padding=10)
        detail_card.pack(fill="x", pady=(8, 0))
        ttk.Label(detail_card, text="文件信息", style="CardSection.TLabel").pack(anchor="w")
        ttk.Label(detail_card, textvariable=self.file_detail_var, style="CardHint.TLabel", wraplength=282, justify="left").pack(anchor="w", pady=(8, 1))
        ttk.Label(
            parent,
            text="单击第一列可启用/停用；选中的第一项作为顶部圈数轴参考文件。",
            style="Hint.TLabel",
            wraplength=280,
        ).pack(anchor="w", pady=(8, 0))

    def _build_plot_panel(self, parent: ttk.Frame) -> None:
        self.figure = Figure(figsize=(10, 7), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, master=parent)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar_frame = ttk.Frame(parent, style="Card.TFrame")
        toolbar_frame.pack(fill="x")
        nav = NavigationToolbar2Tk(self.canvas, toolbar_frame, pack_toolbar=False)
        nav.update()
        nav.pack(side="left", fill="x")

    def _build_settings_panel(self, parent: ttk.Frame) -> None:
        scroll = ScrollableFrame(parent)
        scroll.pack(fill="both", expand=True)
        container = scroll.inner
        ttk.Label(container, text="绘图", style="Section.TLabel").pack(anchor="w", padx=10, pady=(10, 8))

        range_card = ttk.Frame(container, style="Card.TFrame", padding=10)
        range_card.pack(fill="x", padx=8, pady=3)
        ttk.Label(range_card, text="绘图模式", style="CardSection.TLabel").pack(anchor="w")
        mode_combo = ttk.Combobox(
            range_card,
            textvariable=self.plot_mode_var,
            values=tuple(PLOT_MODE_LABELS),
            state="readonly",
        )
        mode_combo.pack(fill="x", pady=(6, 9))
        mode_combo.bind("<<ComboboxSelected>>", self._on_plot_mode_changed)
        ttk.Label(range_card, text="圈数范围", style="CardSection.TLabel").pack(anchor="w")
        range_row = ttk.Frame(range_card, style="Card.TFrame")
        range_row.pack(fill="x", pady=(8, 4))
        ttk.Label(range_row, text="从", style="CardHint.TLabel").pack(side="left")
        ttk.Spinbox(
            range_row,
            from_=1,
            to=1_000_000,
            textvariable=self.plot_cycle_start_var,
            width=8,
        ).pack(side="left", padx=(6, 12))
        ttk.Label(range_row, text="到", style="CardHint.TLabel").pack(side="left")
        ttk.Spinbox(
            range_row,
            from_=1,
            to=1_000_000,
            textvariable=self.plot_cycle_end_var,
            width=8,
        ).pack(side="left", padx=6)
        ttk.Button(range_row, text="默认", command=self._reset_plot_range, width=5).pack(side="right")
        ttk.Label(
            range_card,
            textvariable=self.plot_range_hint_var,
            style="CardHint.TLabel",
            wraplength=300,
        ).pack(anchor="w", pady=(3, 0))
        self.multi_battery_options = ttk.Frame(range_card, style="Card.TFrame")
        ttk.Separator(self.multi_battery_options).pack(fill="x", pady=(9, 7))
        error_row = ttk.Frame(self.multi_battery_options, style="Card.TFrame")
        error_row.pack(fill="x")
        ttk.Label(error_row, text="误差棒", style="CardHint.TLabel").pack(side="left")
        self.error_bar_combo = ttk.Combobox(
            error_row,
            textvariable=self.error_bar_var,
            values=tuple(ERROR_BAR_LABELS),
            state="readonly",
            width=11,
        )
        self.error_bar_combo.pack(side="right")
        self.individual_cells_check = ttk.Checkbutton(
            self.multi_battery_options,
            text="统计图中显示各电池浅色曲线",
            variable=self.show_individual_cells_var,
        )
        self.individual_cells_check.pack(anchor="w", pady=(6, 1))

        data_card = ttk.Frame(container, style="Card.TFrame", padding=10)
        data_card.pack(fill="x", padx=8, pady=5)
        ttk.Label(data_card, text="绘图区数据", style="CardSection.TLabel").pack(anchor="w")
        ttk.Label(data_card, textvariable=self.curve_count_var, style="CardHint.TLabel").pack(anchor="w", pady=(2, 7))
        list_frame = ttk.Frame(data_card, style="Card.TFrame")
        list_frame.pack(fill="x")
        list_scroll = ttk.Scrollbar(list_frame, orient="vertical")
        self.curve_listbox = tk.Listbox(
            list_frame,
            height=7,
            selectmode="extended",
            activestyle="none",
            bg="#FFFFFF",
            fg="#0F172A",
            selectbackground="#DBEAFE",
            selectforeground="#0F172A",
            highlightthickness=1,
            highlightbackground="#CBD5E1",
            relief="flat",
            yscrollcommand=list_scroll.set,
            font=("Segoe UI", 10),
        )
        list_scroll.configure(command=self.curve_listbox.yview)
        self.curve_listbox.pack(side="left", fill="both", expand=True)
        list_scroll.pack(side="right", fill="y")
        self.curve_listbox.bind("<Delete>", lambda _event: self.remove_plot_metrics())
        self.curve_listbox.bind("<<ListboxSelect>>", self._on_curve_list_selected)
        metric_buttons = ttk.Frame(data_card, style="Card.TFrame")
        metric_buttons.pack(fill="x", pady=(8, 0))
        ttk.Button(metric_buttons, text="＋ 添加数据", command=self.add_plot_metrics).pack(side="left")
        ttk.Button(metric_buttons, text="－ 移除", command=self.remove_plot_metrics).pack(side="left", padx=(7, 0))
        self.common_metrics_button = ttk.Button(
            metric_buttons,
            text="常用曲线",
            command=self.add_common_statistics_metrics,
        )
        self.common_metrics_button.pack(side="right")

        ttk.Button(
            container,
            text="绘制所选范围",
            command=self.draw_current_plot,
            style="Accent.TButton",
        ).pack(fill="x", padx=8, pady=(7, 5))

        advanced = CollapsibleFrame(container, "高级设置", initially_open=False)
        advanced.pack(fill="x", padx=8, pady=(5, 18))
        body = advanced.body
        ttk.Label(body, text="指标", style="CardHint.TLabel").grid(row=0, column=0, sticky="w")
        self.style_metric_combo = ttk.Combobox(body, textvariable=self.style_metric_var, state="readonly", width=25)
        self.style_metric_combo.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(2, 8))
        self.style_metric_combo.bind("<<ComboboxSelected>>", self._load_style_form)

        ttk.Label(body, text="颜色", style="CardHint.TLabel").grid(row=2, column=0, sticky="w")
        self.color_button = tk.Button(body, text="     ", relief="flat", command=self._choose_curve_color, cursor="hand2")
        self.color_button.grid(row=3, column=0, sticky="w", pady=(2, 8))
        ttk.Label(body, text="绘制方式", style="CardHint.TLabel").grid(row=2, column=1, sticky="w", padx=(8, 0))
        plot_combo = ttk.Combobox(body, textvariable=self.plot_type_var, state="readonly", values=("线", "点", "线+点"), width=9)
        plot_combo.grid(row=3, column=1, sticky="ew", padx=(8, 0), pady=(2, 8))
        ttk.Label(body, text="轴位置", style="CardHint.TLabel").grid(row=2, column=2, sticky="w", padx=(8, 0))
        side_combo = ttk.Combobox(body, textvariable=self.side_var, state="readonly", values=("自动", "左", "右"), width=8)
        side_combo.grid(row=3, column=2, sticky="ew", padx=(8, 0), pady=(2, 8))

        ttk.Label(body, text="轴外移 (px)", style="CardHint.TLabel").grid(row=4, column=0, sticky="w")
        ttk.Spinbox(body, from_=0, to=240, increment=10, textvariable=self.offset_var, width=8).grid(row=5, column=0, sticky="w", pady=(2, 8))
        ttk.Label(body, text="线宽", style="CardHint.TLabel").grid(row=4, column=1, sticky="w", padx=(8, 0))
        ttk.Spinbox(body, from_=0.5, to=6, increment=0.2, textvariable=self.line_width_var, width=8).grid(row=5, column=1, sticky="ew", padx=(8, 0), pady=(2, 8))
        ttk.Label(body, text="点大小", style="CardHint.TLabel").grid(row=4, column=2, sticky="w", padx=(8, 0))
        ttk.Spinbox(body, from_=1, to=14, increment=0.5, textvariable=self.marker_size_var, width=8).grid(row=5, column=2, sticky="ew", padx=(8, 0), pady=(2, 8))
        ttk.Checkbutton(body, text="显示该指标的 Y 轴", variable=self.axis_visible_var).grid(row=6, column=0, columnspan=2, sticky="w")
        ttk.Button(body, text="应用到当前指标", command=self._apply_style_form, style="Accent.TButton").grid(row=7, column=0, columnspan=3, sticky="ew", pady=(9, 0))

        ttk.Separator(body).grid(row=8, column=0, columnspan=3, sticky="ew", pady=13)
        ttk.Label(body, text="图面与导出", style="CardSection.TLabel").grid(row=9, column=0, columnspan=3, sticky="w", pady=(0, 7))
        ttk.Label(body, text="图标题", style="CardHint.TLabel").grid(row=10, column=0, columnspan=3, sticky="w")
        ttk.Entry(body, textvariable=self.title_var).grid(row=11, column=0, columnspan=3, sticky="ew", pady=(2, 8))
        ttk.Label(body, text="时间单位", style="CardHint.TLabel").grid(row=12, column=0, sticky="w")
        ttk.Combobox(body, textvariable=self.time_unit_var, values=("s", "min", "h"), state="readonly", width=8).grid(row=13, column=0, sticky="w", pady=(2, 8))
        ttk.Label(body, text="每条曲线最大点数", style="CardHint.TLabel").grid(row=12, column=1, columnspan=2, sticky="w", padx=(8, 0))
        ttk.Spinbox(body, from_=1000, to=100000, increment=1000, textvariable=self.max_points_var, width=12).grid(row=13, column=1, columnspan=2, sticky="ew", padx=(8, 0), pady=(2, 8))
        ttk.Checkbutton(body, text="下方时间轴", variable=self.show_bottom_var).grid(row=14, column=0, columnspan=2, sticky="w")
        ttk.Checkbutton(body, text="上方圈数轴", variable=self.show_top_var).grid(row=14, column=2, sticky="w")
        ttk.Checkbutton(body, text="网格", variable=self.show_grid_var).grid(row=15, column=0, columnspan=2, sticky="w")
        ttk.Checkbutton(body, text="图例", variable=self.show_legend_var).grid(row=15, column=2, sticky="w")
        ttk.Label(body, text="导出 DPI", style="CardHint.TLabel").grid(row=16, column=0, sticky="w", pady=(8, 0))
        ttk.Spinbox(body, from_=72, to=1200, increment=50, textvariable=self.dpi_var, width=9).grid(row=17, column=0, sticky="w", pady=(2, 0))
        for column in range(3):
            body.columnconfigure(column, weight=1)

    def _set_busy(self, busy: bool, message: str | None = None) -> None:
        if busy:
            self.progress.start(12)
        else:
            self.progress.stop()
        if message:
            self.status_var.set(message)

    def add_files(self) -> None:
        paths = filedialog.askopenfilenames(
            parent=self,
            title="选择蓝电 CSV 文件",
            filetypes=(("CSV files", "*.csv"), ("All files", "*.*")),
        )
        self._queue_files(paths)

    def add_folder(self) -> None:
        folder = filedialog.askdirectory(parent=self, title="选择包含蓝电 CSV 的文件夹")
        if not folder:
            return
        paths = sorted(Path(folder).glob("*.csv"), key=lambda item: item.name.lower())
        if not paths:
            messagebox.showinfo(APP_NAME, "所选文件夹中没有 CSV 文件。", parent=self)
            return
        self._queue_files(paths)

    def load_paths(self, paths: Sequence[str | Path]) -> None:
        """Public helper used by tests and future drag-and-drop integrations."""
        self._queue_files(paths)

    def _queue_files(self, paths: Sequence[str | Path]) -> None:
        added = 0
        for raw_path in paths:
            path = str(Path(raw_path).resolve())
            if path in self.datasets or path in self.loading_paths:
                continue
            self.loading_paths.add(path)
            self.file_tree.insert("", "end", iid=path, text=f"☑ {Path(path).name}", values=("解析中", "", ""), tags=("loading",))
            future = self.executor.submit(parse_land_file, path)
            future.add_done_callback(lambda result, file_path=path: self.ui_queue.put(("parse", (file_path, result))))
            added += 1
        if added:
            self._set_busy(True, f"正在后台解析 {added} 个文件…")

    def _poll_ui_queue(self) -> None:
        try:
            while True:
                kind, payload = self.ui_queue.get_nowait()
                if kind == "parse":
                    path, future = payload
                    self.loading_paths.discard(path)
                    if not self.file_tree.exists(path):
                        continue
                    try:
                        dataset = future.result()
                    except Exception as exc:
                        self.file_tree.item(path, text=f"☐ {Path(path).name}", values=("错误", "", ""), tags=("error",))
                        self.datasets.pop(path, None)
                        self.status_var.set(f"解析失败：{Path(path).name} — {exc}")
                    else:
                        self.datasets[path] = dataset
                        self.file_tree.item(
                            path,
                            text=f"☑ {dataset.path.name}",
                            values=("就绪", dataset.cycle_count, f"{len(dataset.records):,}"),
                            tags=(),
                        )
                        self._update_range_hint()
                        self.status_var.set(f"已读取 {dataset.path.name}：{dataset.cycle_count} 圈，{len(dataset.records):,} 条记录")
                    if not self.loading_paths:
                        self._set_busy(False)
                        if self.enabled_datasets():
                            self.draw_current_plot()
                elif kind == "export_ok":
                    self._set_busy(False, f"导出完成：{payload}")
                    messagebox.showinfo(APP_NAME, f"导出完成：\n{payload}", parent=self)
                elif kind == "export_error":
                    self._set_busy(False, "导出失败")
                    messagebox.showerror(APP_NAME, f"导出失败：\n{payload}", parent=self)
                elif kind == "status":
                    self.status_var.set(str(payload))
        except queue.Empty:
            pass
        if self.winfo_exists():
            self.after(100, self._poll_ui_queue)

    def enabled_datasets(self) -> list[LandDataset]:
        return [dataset for dataset in self.datasets.values() if dataset.enabled]

    def selected_metric_keys(self) -> list[str]:
        return list(self.plot_metric_keys)

    def _refresh_curve_list(self) -> None:
        self.curve_listbox.delete(0, "end")
        for key in self.plot_metric_keys:
            self.curve_listbox.insert("end", METRIC_BY_KEY[key].display)
        count = len(self.plot_metric_keys)
        self.curve_count_var.set(f"已添加 {count} 项" if count else "尚未添加数据；点击 ＋ 开始")

    def add_plot_metrics(self) -> None:
        source_filter: Literal["record", "cycle"] | None = None
        mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
        if mode == "statistics":
            source_filter = "cycle"
        elif mode == "stack":
            source_filter = "record"
        dialog = AddMetricDialog(self, self.plot_metric_keys, source_filter=source_filter)
        self.wait_window(dialog)
        if not dialog.result:
            return
        self.plot_metric_keys.extend(key for key in dialog.result if key not in self.plot_metric_keys)
        self._refresh_curve_list()
        self._refresh_style_metric_choices()
        self.status_var.set(f"已向绘图区添加 {len(dialog.result)} 项数据；点击“绘制所选范围”。")

    def add_common_statistics_metrics(self) -> None:
        mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
        if mode == "statistics":
            common_keys = ["coulombic_efficiency_pct", "discharge_specific_capacity_mah_g"]
            added_message = "已添加库伦效率和放电比容量；多电池统计模式默认绘制前 20 圈。"
            existing_message = "库伦效率和放电比容量已经在绘图区中。"
        else:
            common_keys = ["voltage_v", "current_ma"]
            added_message = "已添加电压和电流；原始曲线与 Stack 模式默认绘制前 5 圈。"
            existing_message = "电压和电流已经在绘图区中。"
        added = [key for key in common_keys if key not in self.plot_metric_keys]
        self.plot_metric_keys.extend(added)
        self._refresh_curve_list()
        self._refresh_style_metric_choices()
        if added:
            self.status_var.set(added_message)
        else:
            self.status_var.set(existing_message)

    def remove_plot_metrics(self) -> None:
        indices = list(self.curve_listbox.curselection())
        if not indices:
            return
        for index in reversed(indices):
            del self.plot_metric_keys[index]
        self._refresh_curve_list()
        self._refresh_style_metric_choices()
        self.status_var.set(f"绘图区还剩 {len(self.plot_metric_keys)} 项数据。")

    def _on_curve_list_selected(self, _event: tk.Event | None = None) -> None:
        selection = self.curve_listbox.curselection()
        if not selection:
            return
        key = self.plot_metric_keys[selection[0]]
        self.style_metric_var.set(METRIC_BY_KEY[key].display)
        self._load_style_form()

    def _reset_plot_range(self) -> None:
        self.plot_cycle_start_var.set(1)
        mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
        self.plot_cycle_end_var.set(20 if mode == "statistics" else 5)

    def _sync_plot_mode_controls(self) -> None:
        mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
        self.common_metrics_button.configure(text="常用统计" if mode == "statistics" else "电压＋电流")
        if mode != "statistics":
            self.multi_battery_options.pack_forget()
            return
        if not self.multi_battery_options.winfo_manager():
            self.multi_battery_options.pack(fill="x")
        self.error_bar_combo.configure(state="readonly")
        self.individual_cells_check.configure(state="normal")

    def _on_plot_mode_changed(self, _event: tk.Event | None = None) -> None:
        mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
        try:
            current_start = int(self.plot_cycle_start_var.get())
            current_end = int(self.plot_cycle_end_var.get())
        except (tk.TclError, ValueError):
            current_start, current_end = 1, 5
        if mode == "statistics" and current_start == 1 and current_end <= 5:
            self.plot_cycle_end_var.set(20)
        elif mode in {"time", "stack"} and current_start == 1 and current_end == 20:
            self.plot_cycle_end_var.set(5)
        default_titles = {
            "time": "LAND electrochemical data",
            "statistics": "Multi-battery statistics",
            "stack": "Multi-battery raw-curve stack",
        }
        if self.title_var.get() in set(default_titles.values()):
            self.title_var.set(default_titles[mode])
        self._sync_plot_mode_controls()
        if mode == "statistics":
            non_cycle_count = sum(METRIC_BY_KEY[key].source != "cycle" for key in self.plot_metric_keys)
            suffix = f"；当前 {non_cycle_count} 个逐时刻指标将在该模式中忽略" if non_cycle_count else ""
            self.status_var.set(f"多电池统计默认绘制前 20 圈，点击 ＋ 可添加库伦效率、放电比容量等指标{suffix}。")
        elif mode == "stack":
            cycle_count = sum(METRIC_BY_KEY[key].source != "record" for key in self.plot_metric_keys)
            suffix = f"；当前 {cycle_count} 个循环汇总指标将在该模式中忽略" if cycle_count else ""
            self.status_var.set(f"Stack 模式按电池分层比较前 5 圈原始曲线，点击 ＋ 可添加电压、电流等逐时刻指标{suffix}。")
        else:
            self.status_var.set("已切换为原始时序图。")

    def _update_range_hint(self) -> None:
        dataset = self.primary_dataset()
        if not dataset or not dataset.cycles:
            mode = PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time")
            if mode == "statistics":
                hint = "默认统计第 1–20 圈"
            elif mode == "stack":
                hint = "默认分层比较第 1–5 圈"
            else:
                hint = "默认只绘制第 1–5 圈"
            self.plot_range_hint_var.set(hint)
            return
        cycles = sorted(dataset.cycles)
        self.plot_range_hint_var.set(f"当前参考文件：第 {cycles[0]}–{cycles[-1]} 圈，共 {len(cycles)} 圈")

    def primary_dataset(self) -> LandDataset | None:
        for item in self.file_tree.selection():
            dataset = self.datasets.get(item)
            if dataset and dataset.enabled:
                return dataset
        enabled = self.enabled_datasets()
        return enabled[0] if enabled else None

    def _file_tree_click(self, event: tk.Event) -> None:
        row_id = self.file_tree.identify_row(event.y)
        column = self.file_tree.identify_column(event.x)
        if row_id and column == "#0" and row_id in self.datasets:
            dataset = self.datasets[row_id]
            dataset.enabled = not dataset.enabled
            prefix = "☑" if dataset.enabled else "☐"
            self.file_tree.item(row_id, text=f"{prefix} {dataset.path.name}")
            self._update_range_hint()
            self.after_idle(self.draw_current_plot)

    def toggle_selected_files(self) -> None:
        selection = [item for item in self.file_tree.selection() if item in self.datasets]
        if not selection:
            return
        for item in selection:
            dataset = self.datasets[item]
            dataset.enabled = not dataset.enabled
            self.file_tree.item(item, text=f"{'☑' if dataset.enabled else '☐'} {dataset.path.name}")
        self._update_range_hint()
        self.draw_current_plot()

    def remove_selected_files(self) -> None:
        for item in self.file_tree.selection():
            self.datasets.pop(item, None)
            self.loading_paths.discard(item)
            if self.file_tree.exists(item):
                self.file_tree.delete(item)
        self._on_file_selected()
        self.draw_current_plot()

    def clear_files(self) -> None:
        if self.datasets and not messagebox.askyesno(APP_NAME, "清空当前文件列表？", parent=self):
            return
        self.datasets.clear()
        self.loading_paths.clear()
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        self.file_detail_var.set("尚未导入文件")
        self._update_range_hint()
        self.draw_current_plot()

    def _on_file_selected(self, _event: tk.Event | None = None) -> None:
        selection = self.file_tree.selection()
        if not selection:
            self.file_detail_var.set("尚未选择文件")
            self._update_range_hint()
            return
        path = selection[0]
        dataset = self.datasets.get(path)
        if not dataset:
            self.file_detail_var.set(f"{Path(path).name}\n正在解析或解析失败")
            self._update_range_hint()
            return
        duration_h = dataset.duration_s / 3600
        warning_text = f"\n警告：{len(dataset.warnings)} 条" if dataset.warnings else ""
        self.file_detail_var.set(
            f"{dataset.path.name}\n{dataset.cycle_count} 圈 · {len(dataset.steps)} 步 · {len(dataset.records):,} 条记录\n"
            f"总时长 {duration_h:.2f} h{warning_text}\n{dataset.path.parent}"
        )
        self._update_range_hint()

    def _show_file_details(self, _event: tk.Event | None = None) -> None:
        dataset = self.primary_dataset()
        if not dataset:
            return
        warnings = "\n".join(dataset.warnings[:10]) if dataset.warnings else "无解析警告"
        messagebox.showinfo(
            f"{dataset.path.name} — 文件详情",
            f"路径：{dataset.path}\n循环：{dataset.cycle_count}\n步骤：{len(dataset.steps)}\n"
            f"逐时刻记录：{len(dataset.records):,}\n测试时长：{dataset.duration_s/3600:.3f} h\n\n{warnings}",
            parent=self,
        )

    def _metric_display_to_key(self, display: str) -> str | None:
        for metric in METRICS:
            if metric.display == display:
                return metric.key
        return None

    def _refresh_style_metric_choices(self) -> None:
        selected = [METRIC_BY_KEY[key].display for key in self.selected_metric_keys()]
        self.style_metric_combo.configure(values=selected)
        if selected and self.style_metric_var.get() not in selected:
            self.style_metric_var.set(selected[0])
            self._load_style_form()
        elif not selected:
            self.style_metric_var.set("")

    def _load_style_form(self, _event: tk.Event | None = None) -> None:
        key = self._metric_display_to_key(self.style_metric_var.get())
        if not key:
            return
        style = self.curve_styles[key]
        plot_labels = {"line": "线", "point": "点", "line+point": "线+点"}
        side_labels = {"auto": "自动", "left": "左", "right": "右"}
        self.plot_type_var.set(plot_labels[style.plot_type])
        self.side_var.set(side_labels[style.side])
        self.offset_var.set(style.offset)
        self.axis_visible_var.set(style.axis_visible)
        self.line_width_var.set(style.line_width)
        self.marker_size_var.set(style.marker_size)
        self.color_button.configure(bg=style.color, activebackground=style.color)

    def _choose_curve_color(self) -> None:
        key = self._metric_display_to_key(self.style_metric_var.get())
        if not key:
            return
        color = colorchooser.askcolor(self.curve_styles[key].color, parent=self, title="选择曲线颜色")[1]
        if color:
            self.curve_styles[key].color = color
            self.color_button.configure(bg=color, activebackground=color)

    def _apply_style_form(self, redraw: bool = True) -> None:
        key = self._metric_display_to_key(self.style_metric_var.get())
        if not key:
            return
        plot_types = {"线": "line", "点": "point", "线+点": "line+point"}
        sides = {"自动": "auto", "左": "left", "右": "right"}
        style = self.curve_styles[key]
        style.plot_type = plot_types.get(self.plot_type_var.get(), "line")  # type: ignore[assignment]
        style.side = sides.get(self.side_var.get(), "auto")  # type: ignore[assignment]
        style.offset = max(0, int(self.offset_var.get()))
        style.axis_visible = bool(self.axis_visible_var.get())
        style.line_width = max(0.1, float(self.line_width_var.get()))
        style.marker_size = max(0.5, float(self.marker_size_var.get()))
        if redraw:
            self.draw_current_plot()

    def current_plot_options(self) -> PlotOptions:
        cycle_start = int(self.plot_cycle_start_var.get())
        cycle_end = int(self.plot_cycle_end_var.get())
        if cycle_start < 1 or cycle_end < 1 or cycle_start > cycle_end:
            raise ValueError("绘图圈数需为正整数，且起始圈不能大于结束圈。")
        return PlotOptions(
            title=self.title_var.get().strip() or "LAND electrochemical data",
            time_unit=self.time_unit_var.get(),  # type: ignore[arg-type]
            cycle_start=cycle_start,
            cycle_end=cycle_end,
            plot_mode=PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time"),  # type: ignore[arg-type]
            error_bar=ERROR_BAR_LABELS.get(self.error_bar_var.get(), "sd"),  # type: ignore[arg-type]
            show_individual_cells=self.show_individual_cells_var.get(),
            show_bottom_time=self.show_bottom_var.get(),
            show_top_cycle=self.show_top_var.get(),
            show_grid=self.show_grid_var.get(),
            show_legend=self.show_legend_var.get(),
            max_points_per_curve=max(1000, int(self.max_points_var.get())),
        )

    def draw_current_plot(self) -> None:
        try:
            self._apply_style_form(redraw=False)
            datasets = self.enabled_datasets()
            metrics = self.selected_metric_keys()
            options = self.current_plot_options()
            if options.plot_mode == "time" and len(metrics) > 8:
                self.status_var.set("提示：当前绘制超过 8 个 Y 轴，图面可能较拥挤。")
            draw_selected_plot(
                self.figure,
                datasets,
                metrics,
                self.curve_styles,
                options,
                self.primary_dataset(),
            )
            self.canvas.draw_idle()
            if datasets and metrics:
                mode_text = {
                    "time": "原始时序图",
                    "statistics": "多电池均值＋误差棒",
                    "stack": "多电池原始曲线 Stack",
                }[options.plot_mode]
                if options.plot_mode == "statistics":
                    visible_metrics = _cycle_metric_keys(metrics)
                elif options.plot_mode == "stack":
                    visible_metrics = _record_metric_keys(metrics)
                else:
                    visible_metrics = metrics
                self.status_var.set(
                    f"{mode_text} · 第 {options.cycle_start}–{options.cycle_end} 圈 · "
                    f"{len(datasets)} 个电池 × {len(visible_metrics)} 个指标"
                )
        except Exception as exc:
            self.status_var.set(f"绘图失败：{exc}")
            messagebox.showerror(APP_NAME, f"绘图失败：\n{exc}", parent=self)

    def export_data_dialog(self) -> None:
        datasets = self.enabled_datasets()
        if not datasets:
            messagebox.showwarning(APP_NAME, "请先导入并启用至少一个文件。", parent=self)
            return
        try:
            default_start = int(self.plot_cycle_start_var.get())
            default_end = int(self.plot_cycle_end_var.get())
        except (tk.TclError, ValueError):
            default_start, default_end = 1, 5
        default_layout: Literal["wide", "statistics"] = (
            "statistics" if PLOT_MODE_LABELS.get(self.plot_mode_var.get(), "time") == "statistics" else "wide"
        )
        dialog = DataExportDialog(self, default_start, default_end, default_layout=default_layout)
        self.wait_window(dialog)
        if not dialog.result:
            return
        metric_keys = self.selected_metric_keys() if dialog.result["scope"] == "selected" else [metric.key for metric in METRICS]
        if dialog.result["layout"] == "statistics":
            metric_keys = _cycle_metric_keys(metric_keys)
        if not metric_keys:
            messagebox.showwarning(
                APP_NAME,
                "当前选择中没有可导出的指标。多电池统计表需要库伦效率、放电比容量等循环汇总指标。",
                parent=self,
            )
            return
        output = filedialog.asksaveasfilename(
            parent=self,
            title="保存整理后的数据",
            defaultextension=".xlsx",
            filetypes=(("Excel workbook", "*.xlsx"), ("CSV", "*.csv")),
            initialfile="LAND多电池统计.xlsx" if dialog.result["layout"] == "statistics" else "LAND整理数据.xlsx",
        )
        if not output:
            return
        self._set_busy(True, "正在后台导出数据…")

        def run_export() -> str:
            if dialog.result["layout"] == "statistics":
                export_cycle_statistics(
                    datasets,
                    metric_keys,
                    output,
                    cycle_start=dialog.result["cycle_start"],
                    cycle_end=dialog.result["cycle_end"],
                    progress=lambda text: self.ui_queue.put(("status", text)),
                )
            else:
                export_selected_data(
                    datasets,
                    metric_keys,
                    output,
                    layout=dialog.result["layout"],  # type: ignore[arg-type]
                    cycle_start=dialog.result["cycle_start"],
                    cycle_end=dialog.result["cycle_end"],
                    progress=lambda text: self.ui_queue.put(("status", text)),
                )
            return output

        future = self.executor.submit(run_export)

        def finished(result: Any) -> None:
            try:
                self.ui_queue.put(("export_ok", result.result()))
            except Exception as exc:
                self.ui_queue.put(("export_error", str(exc)))

        future.add_done_callback(finished)

    def export_figure_dialog(self) -> None:
        if not self.enabled_datasets() or not self.selected_metric_keys():
            messagebox.showwarning(APP_NAME, "请先绘制至少一个数据指标。", parent=self)
            return
        output = filedialog.asksaveasfilename(
            parent=self,
            title="导出当前图像",
            defaultextension=".png",
            filetypes=(
                ("PNG image", "*.png"),
                ("SVG vector", "*.svg"),
                ("PDF vector", "*.pdf"),
                ("Enhanced Metafile", "*.emf"),
                ("EPS vector", "*.eps"),
                ("TIFF image", "*.tif;*.tiff"),
            ),
            initialfile="LAND_plot.png",
        )
        if not output:
            return
        try:
            save_figure(self.figure, output, dpi=max(72, int(self.dpi_var.get())))
        except Exception as exc:
            messagebox.showerror(APP_NAME, f"图像导出失败：\n{exc}", parent=self)
            return
        self.status_var.set(f"图像已导出：{output}")
        messagebox.showinfo(APP_NAME, f"图像已导出：\n{output}", parent=self)

    def _on_close(self) -> None:
        self.executor.shutdown(wait=False, cancel_futures=True)
        self.destroy()


def make_test_subset(dataset: LandDataset, record_limit: int = 3000, cycle_limit: int = 20) -> LandDataset:
    cycles = dict(list(sorted(dataset.cycles.items()))[:cycle_limit])
    allowed_cycles = set(cycles)
    records = [row for row in dataset.records if row.cycle in allowed_cycles][:record_limit]
    allowed_step_keys = {(row.cycle, row.step) for row in records}
    steps = {key: value for key, value in dataset.steps.items() if key in allowed_step_keys}
    subset = LandDataset(
        path=dataset.path,
        name=dataset.name,
        records=records,
        steps=steps,
        cycles=cycles,
        warnings=list(dataset.warnings),
    )
    subset.rebuild_index()
    return subset


def run_self_test(source_dir: str | Path, output_dir: str | Path) -> int:
    source = Path(source_dir)
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    csv_files = sorted(source.glob("*.csv"), key=lambda item: item.name.lower())
    if not csv_files:
        raise FileNotFoundError(f"没有找到 CSV：{source}")

    selected_files: list[Path] = []
    for prefix in ("c", "f"):
        candidate = next((item for item in csv_files if item.name.lower().startswith(prefix)), None)
        if candidate:
            selected_files.append(candidate)
    if not selected_files:
        selected_files = csv_files[:1]

    parsed = [parse_land_file(path) for path in selected_files]
    subsets = [make_test_subset(dataset) for dataset in parsed]
    for dataset in parsed:
        if not dataset.records or not dataset.cycles:
            raise AssertionError(f"解析结果为空：{dataset.path}")
        first = dataset.records[0]
        if first.cycle < 1 or first.step < 1:
            raise AssertionError(f"循环/步骤编号异常：{dataset.path}")

    metric_keys = ["voltage_v", "current_ma", "coulombic_efficiency_pct"]
    styles = {metric.key: CurveStyle(metric.default_color, metric.default_style) for metric in METRICS}
    figure = Figure(figsize=(11, 6), dpi=110)
    draw_plot(
        figure,
        subsets,
        metric_keys,
        styles,
        PlotOptions(title="LAND Workbench self-test", max_points_per_curve=3000),
        subsets[0],
    )
    save_figure(figure, output / "self_test_plot.png", dpi=160)
    save_figure(figure, output / "self_test_plot.svg", dpi=160)
    if os.name == "nt":
        save_figure(figure, output / "self_test_plot.emf", dpi=160)
    export_selected_data(
        subsets,
        metric_keys,
        output / "self_test_wide.csv",
        layout="wide",
        cycle_start=1,
        cycle_end=5,
    )

    statistics_keys = ["coulombic_efficiency_pct", "discharge_capacity_mah"]
    statistics = compute_cycle_statistics(subsets, statistics_keys[0], 1, 20)
    if not np.any(statistics.n > 0):
        raise AssertionError("多电池统计结果为空。")
    statistics_figure = Figure(figsize=(10, 7), dpi=110)
    draw_cycle_statistics_plot(
        statistics_figure,
        subsets,
        statistics_keys,
        styles,
        PlotOptions(
            title="LAND multi-battery statistics self-test",
            cycle_start=1,
            cycle_end=20,
            plot_mode="statistics",
            error_bar="sd",
        ),
    )
    save_figure(statistics_figure, output / "self_test_statistics.png", dpi=160)
    stack_figure = Figure(figsize=(10, 7), dpi=110)
    draw_record_stack_plot(
        stack_figure,
        subsets,
        ["voltage_v", "current_ma"],
        styles,
        PlotOptions(title="LAND raw-curve stack self-test", cycle_start=1, cycle_end=5, plot_mode="stack"),
    )
    save_figure(stack_figure, output / "self_test_stack.png", dpi=160)
    export_cycle_statistics(
        subsets,
        statistics_keys,
        output / "self_test_statistics.xlsx",
        cycle_start=1,
        cycle_end=20,
    )
    export_selected_data(
        subsets,
        metric_keys,
        output / "self_test_long.xlsx",
        layout="long",
        cycle_start=1,
        cycle_end=5,
    )

    report = {
        "version": APP_VERSION,
        "files": [
            {
                "name": dataset.path.name,
                "cycles": dataset.cycle_count,
                "steps": len(dataset.steps),
                "records": len(dataset.records),
                "duration_h": dataset.duration_s / 3600,
                "warnings": len(dataset.warnings),
            }
            for dataset in parsed
        ],
        "outputs": sorted(item.name for item in output.iterdir()),
    }
    (output / "self_test_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


def expand_startup_paths(values: Sequence[str]) -> list[Path]:
    result: list[Path] = []
    for value in values:
        path = Path(value)
        if path.is_dir():
            result.extend(sorted(path.glob("*.csv"), key=lambda item: item.name.lower()))
        elif path.is_file():
            result.append(path)
    return result


def create_gui_snapshot(csv_path: str | Path, output_path: str | Path) -> None:
    try:
        from PIL import ImageGrab
    except ImportError as exc:
        raise RuntimeError("GUI 截图测试需要 Pillow。") from exc
    dataset = parse_land_file(csv_path)
    app = LandtWorkbenchApp()
    app.geometry("1500x900+40+40")
    key = str(dataset.path.resolve())
    app.datasets[key] = dataset
    app.file_tree.insert(
        "",
        "end",
        iid=key,
        text=f"☑ {dataset.path.name}",
        values=("就绪", dataset.cycle_count, f"{len(dataset.records):,}"),
    )
    app.file_tree.selection_set(key)
    app._on_file_selected()
    app.plot_metric_keys = ["voltage_v", "current_ma", "coulombic_efficiency_pct"]
    app._refresh_curve_list()
    app._refresh_style_metric_choices()
    app.draw_current_plot()
    app.update_idletasks()
    app.attributes("-topmost", True)
    app.lift()
    app.update()
    time.sleep(0.5)
    if os.name == "nt":
        rect = RECT()
        ctypes.windll.user32.GetWindowRect(app.winfo_id(), ctypes.byref(rect))
        x1, y1, x2, y2 = rect.left, rect.top, rect.right, rect.bottom
    else:
        x1, y1 = app.winfo_rootx(), app.winfo_rooty()
        x2, y2 = x1 + app.winfo_width(), y1 + app.winfo_height()
    ImageGrab.grab(bbox=(x1, y1, x2, y2), all_screens=True).save(output_path)
    app.attributes("-topmost", False)
    app.destroy()


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=f"{APP_NAME} {APP_VERSION}")
    parser.add_argument("files", nargs="*", help="启动时自动导入的 CSV 文件或文件夹")
    parser.add_argument("--self-test", metavar="CSV_DIR", help="运行解析/绘图/导出自检，不启动 GUI")
    parser.add_argument("--test-output", default="self_test_output", help="自检输出目录")
    parser.add_argument("--smoke-gui", action="store_true", help="创建并关闭 GUI，用于启动冒烟测试")
    parser.add_argument("--snapshot", metavar="PNG", help="用第一个启动文件生成 GUI 截图并退出")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_argument_parser().parse_args(argv)
    if args.self_test:
        return run_self_test(args.self_test, args.test_output)
    if args.snapshot:
        startup_paths = expand_startup_paths(args.files)
        if not startup_paths:
            raise ValueError("--snapshot 需要提供至少一个 CSV 文件或文件夹。")
        create_gui_snapshot(startup_paths[0], args.snapshot)
        return 0
    app = LandtWorkbenchApp()
    startup_paths = expand_startup_paths(args.files)
    if startup_paths:
        app.after(150, lambda: app.load_paths(startup_paths))
    if args.smoke_gui:
        app.withdraw()
        app.update_idletasks()
        app.update()
        app.destroy()
        return 0
    app.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
